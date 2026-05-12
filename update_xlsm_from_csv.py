# update_colorimetric_report.py

from pathlib import Path
from shutil import copyfile
import re
import unicodedata
import warnings
from datetime import datetime

from openpyxl import load_workbook
from openpyxl.drawing.image import Image


warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")


# ============================================================
# Config
# ============================================================

INPUT_XLSM = "template/F-82 Reporte de Resultados Colorimetricos Ver.05 Bray.xlsm"
OUTPUT_XLSM = "output_updated.xlsm"

CSV_FILE = "input/Cuantificación_23_03_2026_12_52_11.csv"

TARGET_SHEET = "P_DIS"

DATE_CELL = "T12"

# Excel columns in P_DIS:
# F = Concentración (mg/L)
# H = Absorbancia / A882
CONCENTRATION_COLUMN = "F"
A882_COLUMN = "H"

START_ROW = 15

# Use images directly from the extracted folder
IMAGE_DIR = "extracted_images_from_xlsm/media"


# ============================================================
# Helpers
# ============================================================

def norm(text):
    """
    Normalize text so accents and bad encoding do not break matching.
    """
    if text is None:
        return ""

    text = str(text).strip()
    text = text.replace("\x00", "")
    text = text.replace("??", "")
    text = text.replace("?", "A")

    text = unicodedata.normalize("NFKD", text)
    text = "".join(c for c in text if not unicodedata.combining(c))

    text = text.upper()
    text = re.sub(r"\s+", " ", text)

    return text.strip()


def parse_decimal(value):
    """
    Converts:
        0,05450 -> 0.05450
        1,00000 -> 1.00000
    """
    if value is None:
        return None

    value = str(value).strip()
    value = value.replace("\x00", "")

    if value == "":
        return None

    value = value.replace(",", ".")

    try:
        return float(value)
    except ValueError:
        return None


def read_text_auto_encoding(csv_file):
    """
    Reads CSV text trying common encodings.

    Your CSV looks like UTF-16 because Python showed:
        N\x00o\x00m\x00b\x00r\x00e
    """
    csv_file = Path(csv_file)
    raw = csv_file.read_bytes()

    encodings = [
        "utf-16",
        "utf-16-le",
        "utf-16-be",
        "utf-8-sig",
        "latin-1",
    ]

    for enc in encodings:
        try:
            text = raw.decode(enc)

            # Good decoding should not have many null bytes
            if text.count("\x00") < 5:
                print(f"Detected CSV encoding: {enc}")
                return text

        except UnicodeDecodeError:
            pass

    print("WARNING: Could not detect CSV encoding cleanly. Removing null bytes.")
    return raw.decode("latin-1", errors="ignore").replace("\x00", "")


def read_csv_rows(csv_file):
    """
    Reads the CSV as rows.

    Handles:
        - UTF-16
        - tab-separated
        - semicolon-separated
        - multiple-spaces-separated

    Important:
        We do not split by comma because decimals use comma.
    """
    text = read_text_auto_encoding(csv_file)
    text = text.replace("\x00", "")
    text = text.replace("??", "")

    rows = []

    for line in text.splitlines():
        line = line.strip()

        if not line:
            continue

        if "\t" in line:
            parts = line.split("\t")
        elif ";" in line:
            parts = line.split(";")
        elif re.search(r"\s{2,}", line):
            parts = re.split(r"\s{2,}", line)
        else:
            parts = [line]

        parts = [p.strip() for p in parts if p.strip() != ""]

        if parts:
            rows.append(parts)

    print("\nCSV rows detected:")
    for row in rows[:12]:
        print(row)

    return rows


def find_header_indices(rows):
    """
    Finds the CSV columns:

        Nombre de muestra
        A882
        Concentración

    Returns:
        sample_idx, a882_idx, concentration_idx
    """
    for row in rows:
        normalized = [norm(cell) for cell in row]

        sample_idx = None
        a882_idx = None
        concentration_idx = None

        for i, cell in enumerate(normalized):
            if "NOMBRE" in cell and "MUESTRA" in cell:
                sample_idx = i

            if "A882" in cell or cell == "882":
                a882_idx = i

            if "CONCENTRACION" in cell or cell.startswith("CONC"):
                concentration_idx = i

        if a882_idx is not None and concentration_idx is not None:
            if sample_idx is None:
                sample_idx = 0

            print("\nDetected CSV columns:")
            print(f"  Nombre de muestra column: {sample_idx}")
            print(f"  A882 column: {a882_idx}")
            print(f"  Concentración column: {concentration_idx}")

            return sample_idx, a882_idx, concentration_idx

    raise ValueError("Could not find A882 and Concentración headers in CSV.")


def extract_standard_values(csv_file):
    """
    Extracts Estándar 1 to Estándar 7.

    CSV example:

        Nombre de muestra    A882       Concentración
        Estándar 1           0,00000    0,00000
        Estándar 2           0,05450    0,10000
        ...
    """
    rows = read_csv_rows(csv_file)

    sample_idx, a882_idx, concentration_idx = find_header_indices(rows)

    standards = {}

    for row in rows:
        if len(row) <= max(sample_idx, a882_idx, concentration_idx):
            continue

        sample_name = norm(row[sample_idx])

        # Matches:
        #   ESTANDAR 1
        #   ESTÁNDAR 1
        #   EST?NDAR 1
        match = re.search(r"EST[A-Z]*NDAR\s*(\d+)", sample_name)

        if not match:
            continue

        standard_number = int(match.group(1))

        if 1 <= standard_number <= 7:
            a882_value = parse_decimal(row[a882_idx])
            concentration_value = parse_decimal(row[concentration_idx])

            if a882_value is None or concentration_value is None:
                print(f"WARNING: Could not parse row: {row}")
                continue

            standards[standard_number] = {
                "A882": a882_value,
                "Concentracion": concentration_value,
            }

            print(
                f"Found Estándar {standard_number}: "
                f"A882={a882_value}, Concentración={concentration_value}"
            )

    missing = [i for i in range(1, 8) if i not in standards]

    if missing:
        raise ValueError(f"Missing standards in CSV: {missing}")

    return standards


def extract_date_from_filename(csv_file):
    """
    Extracts date from filename like:

        Cuantificación_23_03_2026_12_52_11.csv

    Returns:
        datetime.date
    """
    name = Path(csv_file).stem

    match = re.search(r"(\d{2})_(\d{2})_(\d{4})", name)

    if not match:
        raise ValueError(f"Could not extract date from filename: {csv_file}")

    day, month, year = match.groups()

    return datetime.strptime(f"{day}/{month}/{year}", "%d/%m/%Y").date()


# ============================================================
# Image insertion
# ============================================================

def reinsert_images(wb, image_dir):
    """
    Reinsert fixed images directly from:

        extracted_images_from_xlsm/media

    Known images:
        image2.jpeg = INIA logo
        image3.png  = LABSAF logo
        image4.png  = equation

    This follows the same fixed-image logic from your previous working builder:
    clear old images, then insert known images at known anchors.
    """
    image_dir = Path(image_dir)

    if not image_dir.exists():
        print(f"Image directory not found. Skipping image insertion: {image_dir}")
        return

    images_by_sheet = {
        "P_DIS": [
            # Top-left logos
            ("image2.jpeg", "A1", 0.23),   # INIA logo, top
            ("image3.png", "A3", 0.33),    # LABSAF logo, bottom

            # Equation image
            # Adjust anchor/scale if needed
            ("image4.png", "N19", 0.55),
        ],

        "Resultados": [
            ("image2.jpeg", "A1", 0.23),
            ("image3.png", "A3", 0.33),
        ],

        "Datos": [
            ("image2.jpeg", "A1", 0.23),
            ("image3.png", "A3", 0.33),
        ],
    }

    for sheet_name, images in images_by_sheet.items():
        if sheet_name not in wb.sheetnames:
            print(f"Sheet not found for image insertion: {sheet_name}")
            continue

        ws = wb[sheet_name]

        # Avoid duplicated or broken images
        ws._images = []

        for image_filename, anchor, scale in images:
            image_path = image_dir / image_filename

            if not image_path.exists():
                print(f"Image file not found: {image_path}")
                continue

            img = Image(str(image_path))
            img.width = img.width * scale
            img.height = img.height * scale

            ws.add_image(img, anchor)

            print(
                f"Inserted {image_filename} into {sheet_name}!{anchor} "
                f"at {int(scale * 100)}% size"
            )


# ============================================================
# Excel writing
# ============================================================

def write_standard_values(ws, standards):
    """
    Writes:

        CSV Concentración -> Excel F15:F21
        CSV A882          -> Excel H15:H21
    """
    for standard_number in range(1, 8):
        row = START_ROW + standard_number - 1

        concentration = standards[standard_number]["Concentracion"]
        a882 = standards[standard_number]["A882"]

        ws[f"{CONCENTRATION_COLUMN}{row}"] = concentration
        ws[f"{A882_COLUMN}{row}"] = a882

        ws[f"{CONCENTRATION_COLUMN}{row}"].number_format = "0.00000"
        ws[f"{A882_COLUMN}{row}"].number_format = "0.00000"


def update_report(input_xlsm, output_xlsm, csv_file, image_dir):
    input_xlsm = Path(input_xlsm).resolve()
    output_xlsm = Path(output_xlsm).resolve()
    csv_file = Path(csv_file).resolve()
    image_dir = Path(image_dir).resolve()

    if not input_xlsm.exists():
        raise FileNotFoundError(f"Template not found: {input_xlsm}")

    if not csv_file.exists():
        raise FileNotFoundError(f"CSV not found: {csv_file}")

    if not image_dir.exists():
        raise FileNotFoundError(f"Image directory not found: {image_dir}")

    output_xlsm.parent.mkdir(parents=True, exist_ok=True)

    print("\nUpdating colorimetric report")
    print(f"Template: {input_xlsm}")
    print(f"CSV: {csv_file}")
    print(f"Images: {image_dir}")
    print(f"Output: {output_xlsm}")

    # Copy template first
    copyfile(input_xlsm, output_xlsm)

    # Read CSV data
    date_value = extract_date_from_filename(csv_file)
    standards = extract_standard_values(csv_file)

    # Open copied XLSM with macros preserved
    # This does not open Excel, so there is no macro popup
    wb = load_workbook(output_xlsm, keep_vba=True)

    if TARGET_SHEET not in wb.sheetnames:
        print("Available sheets:")
        for sheet in wb.sheetnames:
            print(repr(sheet))

        raise KeyError(f"Target sheet not found: {TARGET_SHEET}")

    ws = wb[TARGET_SHEET]

    # Date
    ws[DATE_CELL] = date_value
    ws[DATE_CELL].number_format = "dd/mm/yyyy"

    # Standards
    write_standard_values(ws, standards)

    # Reinsert fixed images
    reinsert_images(wb, image_dir=image_dir)

    # Save
    wb.save(output_xlsm)

    print("\nReport created successfully:")
    print(f"  {output_xlsm}")

    print("\nValues written:")
    print(f"  {TARGET_SHEET}!{DATE_CELL} = {date_value.strftime('%d/%m/%Y')}")

    for standard_number in range(1, 8):
        row = START_ROW + standard_number - 1

        print(
            f"  Estándar {standard_number}: "
            f"{CONCENTRATION_COLUMN}{row} = "
            f"{standards[standard_number]['Concentracion']:.5f}, "
            f"{A882_COLUMN}{row} = "
            f"{standards[standard_number]['A882']:.5f}"
        )


def main():
    update_report(
        input_xlsm=INPUT_XLSM,
        output_xlsm=OUTPUT_XLSM,
        csv_file=CSV_FILE,
        image_dir=IMAGE_DIR,
    )


if __name__ == "__main__":
    main()
