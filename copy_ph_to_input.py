# copy_ph_no_marker_use_max_duplicate.py

from pathlib import Path
import re
import shutil
from openpyxl import load_workbook


# ============================================================
# CONFIGURATION
# ============================================================

INPUT_XLSM = r"G:\Mi unidad\LABSAF ILLPA\input.xlsm"

PH_FOLDER = r"G:\Mi unidad\LABSAF ILLPA\1. Documentos Internos\7.5 Registros Tecnicos\2026\SUELOS\1.pH"

OUTPUT_XLSM = r"G:\Mi unidad\LABSAF ILLPA\input_WITH_PH.xlsm"

# Input file tab
INPUT_SHEET_NAME = "P_DIS"

# pH file tab
PH_SHEET_NAME = "F-103"

# Input file structure
# Now we start at C37, not C36.
INPUT_FIRST_ROW = 37

# Each useful block now has 20 sample rows:
# 37 to 56
# 60 to 79
# 83 to 102
# etc.
BLOCK_SIZE = 20

# After each block, skip 3 rows:
# after 56, next is 60
# after 79, next is 83
# after 102, next is 106
GAP_ROWS = 3

INPUT_CODE_COL = "C"
INPUT_OUTPUT_COL = "E"

# pH file structure
PH_FIRST_ROW = 27
PH_CODE_COL = "C"
PH_VALUE_COL = "H"

VERBOSE = True


# ============================================================
# VERBOSE PRINT
# ============================================================

def vprint(message):
    if VERBOSE:
        print(message)


# ============================================================
# NORMALIZATION FUNCTIONS
# ============================================================

def normalize_text(value):
    """
    Clean Excel text.

    Handles:
        'SU1149-ILL-26
        SU1149-ILL-26
        spaces
        non-breaking spaces
    """
    if value is None:
        return ""

    text = str(value).strip()
    text = text.replace("\xa0", " ")
    text = re.sub(r"\s+", " ", text)

    # Remove Excel leading apostrophe
    text = text.lstrip("'").strip()

    return text


def normalize_code(value):
    """
    Normalize SU code.

    Example:
        'SU1149-ILL-26 -> SU1149-ILL-26
    """
    return normalize_text(value).upper()


def extract_su_number(value):
    """
    Extract SU number.

    Examples:
        SU1149-ILL-26   -> 1149
        'SU1149-ILL-26  -> 1149
        SU0079          -> 79
    """
    text = normalize_text(value).upper()

    match = re.search(r"SU\s*0*(\d+)", text)

    if not match:
        return None

    return int(match.group(1))


def parse_float(value):
    """
    Convert pH value to float.

    Handles:
        7.5
        "7.5"
        "7,5"

    Returns None if conversion fails.
    """
    if value is None:
        return None

    text = normalize_text(value)

    if text == "":
        return None

    text = text.replace(",", ".")

    try:
        return float(text)
    except ValueError:
        return None


# ============================================================
# FILE RANGE FUNCTIONS
# ============================================================

def extract_su_ranges_from_filename(filename):
    """
    Extract SU ranges from filenames.

    Examples:
        SU0668-0767.xlsx
        SU0597-0602- SU0608-0667.xlsx
        SU1144-1175.xlsx

    Returns:
        [(668, 767)]
        [(597, 602), (608, 667)]
        [(1144, 1175)]
    """
    text = normalize_text(filename).upper()

    pattern = r"SU\s*0*(\d+)\s*-\s*0*(\d+)"

    ranges = []

    for start, end in re.findall(pattern, text):
        start = int(start)
        end = int(end)

        if start <= end:
            ranges.append((start, end))
        else:
            ranges.append((end, start))

    return ranges


def find_probable_ph_files(ph_folder, su_number):
    """
    Find pH files whose filename range includes the SU number.
    """
    ph_folder = Path(ph_folder)

    if not ph_folder.exists():
        raise FileNotFoundError(f"pH folder not found: {ph_folder}")

    excel_files = sorted(
        list(ph_folder.glob("*.xlsx")) +
        list(ph_folder.glob("*.xlsm"))
    )

    candidates = []

    print("")
    print("Looking for probable pH file by filename range...")
    print(f"  SU number: SU{su_number}")
    print(f"  Folder: {ph_folder}")
    print(f"  Excel files found: {len(excel_files)}")

    for excel_file in excel_files:
        ranges = extract_su_ranges_from_filename(excel_file.name)

        if not ranges:
            continue

        for start, end in ranges:
            if start <= su_number <= end:
                print("")
                print("  Candidate found:")
                print(f"    File: {excel_file.name}")
                print(f"    Range: SU{start} to SU{end}")
                print(f"    SU{su_number} is inside this range.")
                candidates.append(excel_file)

    return candidates


# ============================================================
# INPUT ROW ITERATOR
# ============================================================

def iter_input_rows(ws):
    """
    Generate input rows in this pattern:

        37 to 56
        60 to 79
        83 to 102
        106 to 125
        ...

    This ignores the D rows:
        36, 59, 82, 105, ...

    We only read the 20 sample rows after each D row.
    """
    row = INPUT_FIRST_ROW

    while row <= ws.max_row:
        block_start = row
        block_end = row + BLOCK_SIZE - 1

        vprint("")
        vprint("------------------------------------------------------------")
        vprint(f"Input block: rows {block_start} to {block_end}")
        vprint("------------------------------------------------------------")

        for r in range(block_start, min(block_end, ws.max_row) + 1):
            yield r

        row = block_end + GAP_ROWS + 1


# ============================================================
# SHEET HELPER
# ============================================================

def get_sheet(workbook, sheet_name):
    if sheet_name not in workbook.sheetnames:
        raise ValueError(
            f"Sheet '{sheet_name}' not found. "
            f"Available sheets: {workbook.sheetnames}"
        )

    return workbook[sheet_name]


# ============================================================
# SEARCH pH FILE
# ============================================================

def search_su_in_ph_file(ph_file, input_code, input_su_number):
    """
    Search one pH file for the SU code.

    It ignores column B completely.

    If the SU appears multiple times, it returns the largest pH value.
    """
    print("")
    print("Opening probable pH file:")
    print(f"  {ph_file}")

    wb = load_workbook(ph_file, data_only=True, read_only=True)

    if PH_SHEET_NAME not in wb.sheetnames:
        wb.close()
        raise ValueError(
            f"Sheet '{PH_SHEET_NAME}' not found in file {ph_file.name}. "
            f"Available sheets: {wb.sheetnames}"
        )

    ws = wb[PH_SHEET_NAME]

    print(f"Accessing pH sheet: {PH_SHEET_NAME}")
    print(f"Searching from row {PH_FIRST_ROW} to {ws.max_row}")
    print(f"Looking for input code: {input_code}")
    print(f"Looking for SU number: {input_su_number}")
    print("Ignoring column B completely.")

    matches = []

    for row in range(PH_FIRST_ROW, ws.max_row + 1):
        raw_code = ws[f"{PH_CODE_COL}{row}"].value
        raw_ph = ws[f"{PH_VALUE_COL}{row}"].value

        ph_code = normalize_code(raw_code)
        ph_su_number = extract_su_number(ph_code)
        ph_value_float = parse_float(raw_ph)

        if VERBOSE:
            print(
                f"  Row {row}: "
                f"C={raw_code!r}->{ph_code!r}, "
                f"SU={ph_su_number}, "
                f"H={raw_ph!r}->{ph_value_float}"
            )

        if ph_su_number is None:
            continue

        if ph_su_number != input_su_number:
            continue

        print("")
        print("  SU MATCH FOUND:")
        print(f"    pH row: {row}")
        print(f"    pH code C: {ph_code}")
        print(f"    pH value H: {raw_ph}")

        if ph_value_float is None:
            print("    WARNING: pH value could not be converted to number. Skipping this match.")
            continue

        matches.append(
            {
                "row": row,
                "code": ph_code,
                "ph_raw": raw_ph,
                "ph_float": ph_value_float,
                "source_file": ph_file.name,
            }
        )

    wb.close()

    if not matches:
        print("")
        print("No usable pH match found in this file.")
        return None

    print("")
    print("Matches found:")
    for match in matches:
        print(
            f"  Row {match['row']} | "
            f"C={match['code']} | "
            f"H={match['ph_raw']} | "
            f"numeric={match['ph_float']}"
        )

    # If duplicated, use the biggest pH value.
    best_match = max(matches, key=lambda m: m["ph_float"])

    print("")
    print("Selected match:")
    print("  Rule: if duplicated, use the biggest pH value.")
    print(f"  Selected row: {best_match['row']}")
    print(f"  Selected pH: {best_match['ph_float']}")
    print(f"  Selected source file: {best_match['source_file']}")

    return best_match


def find_ph_for_input_code(input_code):
    """
    Find pH value for one input SU code.
    """
    input_code_norm = normalize_code(input_code)
    input_su_number = extract_su_number(input_code_norm)

    if input_code_norm == "":
        return None

    if input_su_number is None:
        print(f"Could not extract SU number from input code: {input_code!r}")
        return None

    print("")
    print("============================================================")
    print("SEARCHING PH FOR INPUT CODE")
    print("============================================================")
    print(f"Input code: {input_code!r}")
    print(f"Normalized input code: {input_code_norm}")
    print(f"Extracted SU number: {input_su_number}")

    candidates = find_probable_ph_files(PH_FOLDER, input_su_number)

    if not candidates:
        print("")
        print(f"No probable pH file found for SU{input_su_number}.")
        return None

    all_matches = []

    for candidate in candidates:
        match = search_su_in_ph_file(
            ph_file=candidate,
            input_code=input_code_norm,
            input_su_number=input_su_number,
        )

        if match is not None:
            all_matches.append(match)

    if not all_matches:
        print("")
        print(f"No pH value found inside candidate file(s) for SU{input_su_number}.")
        return None

    # If for some reason the SU appears in more than one candidate file,
    # also select the biggest pH value across all matches.
    best_match = max(all_matches, key=lambda m: m["ph_float"])

    print("")
    print("FINAL SELECTED pH FOR THIS INPUT CODE")
    print(f"  Input code: {input_code_norm}")
    print(f"  Selected pH: {best_match['ph_float']}")
    print(f"  Source file: {best_match['source_file']}")
    print(f"  Source row: {best_match['row']}")

    return best_match


# ============================================================
# MAIN FILL FUNCTION
# ============================================================

def main():
    input_path = Path(INPUT_XLSM)
    output_path = Path(OUTPUT_XLSM)

    print("")
    print("============================================================")
    print("STARTING pH COPY PROCESS WITHOUT COLUMN B / D")
    print("============================================================")
    print(f"Input file:")
    print(f"  {input_path}")
    print(f"Input sheet:")
    print(f"  {INPUT_SHEET_NAME}")
    print(f"pH folder:")
    print(f"  {PH_FOLDER}")
    print(f"pH sheet:")
    print(f"  {PH_SHEET_NAME}")
    print(f"Output file:")
    print(f"  {output_path}")
    print("============================================================")

    if not input_path.exists():
        raise FileNotFoundError(f"Input file not found: {input_path}")

    print("")
    print("Copying input XLSM to output XLSM...")
    shutil.copy2(input_path, output_path)
    print("Copy done.")

    print("")
    print("Opening input workbook for reading values...")
    wb_values = load_workbook(input_path, data_only=True, keep_vba=True)
    ws_values = get_sheet(wb_values, INPUT_SHEET_NAME)

    print(f"Input workbook opened.")
    print(f"Input sheet used: {ws_values.title}")
    print(f"Input max row: {ws_values.max_row}")

    print("")
    print("Opening output workbook for writing values...")
    wb_out = load_workbook(output_path, keep_vba=True)
    ws_out = get_sheet(wb_out, INPUT_SHEET_NAME)

    log = []

    for row in iter_input_rows(ws_values):
        print("")
        print("============================================================")
        print(f"INPUT ROW {row}")
        print("============================================================")

        input_cell = f"{INPUT_CODE_COL}{row}"
        output_cell = f"{INPUT_OUTPUT_COL}{row}"

        raw_code = ws_values[input_cell].value
        code = normalize_code(raw_code)
        su_number = extract_su_number(code)

        print(f"Reading input cell {input_cell}:")
        print(f"  Raw value: {raw_code!r}")
        print(f"  Normalized code: {code!r}")
        print(f"  Extracted SU number: {su_number}")

        if code == "":
            print("Empty code. Skipping.")
            continue

        if su_number is None:
            print("Could not extract SU number. Skipping.")
            log.append(
                {
                    "row": row,
                    "code": code,
                    "ph": "",
                    "status": "SKIPPED: no SU number",
                    "source_file": "",
                    "source_row": "",
                }
            )
            continue

        match = find_ph_for_input_code(code)

        if match is None:
            print("")
            print("No pH found for this input row.")
            log.append(
                {
                    "row": row,
                    "code": code,
                    "ph": "",
                    "status": "NOT FOUND",
                    "source_file": "",
                    "source_row": "",
                }
            )
            continue

        ph_value = match["ph_float"]

        print("")
        print("Writing pH value to output workbook:")
        print(f"  Destination cell: {output_cell}")
        print(f"  Value: {ph_value}")

        ws_out[output_cell] = ph_value

        log.append(
            {
                "row": row,
                "code": code,
                "ph": ph_value,
                "status": "OK",
                "source_file": match["source_file"],
                "source_row": match["row"],
            }
        )

    print("")
    print("Saving output workbook...")
    wb_out.save(output_path)

    wb_values.close()
    wb_out.close()

    print("")
    print("============================================================")
    print("FINISHED")
    print("============================================================")
    print(f"Output file created:")
    print(f"  {output_path}")

    print("")
    print("FINAL SUMMARY")
    print("============================================================")

    for item in log:
        print(
            f"Row {item['row']:>4} | "
            f"C={item['code']:<25} | "
            f"pH={str(item['ph']):<10} | "
            f"{item['status']:<12} | "
            f"{item['source_file']} | "
            f"source row={item['source_row']}"
        )


if __name__ == "__main__":
    main()
