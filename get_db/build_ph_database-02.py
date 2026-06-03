from pathlib import Path
import re
import csv
import sqlite3
import argparse
import subprocess
import tempfile
import io
import hashlib

from openpyxl import load_workbook


try:
    import msoffcrypto
except ImportError:
    msoffcrypto = None


# ============================================================
# PRINT / LOG HELPER
# ============================================================

def default_log(message):
    print(message)


# ============================================================
# NORMALIZATION
# ============================================================

def normalize_text(value):
    if value is None:
        return ""

    text = str(value).strip()
    text = text.replace("\xa0", " ")
    text = re.sub(r"\s+", " ", text)

    # Ignore leading Excel apostrophe:
    # 'SU1149-ILL-26 -> SU1149-ILL-26
    text = text.lstrip("'").strip()

    return text


def normalize_code(value):
    return normalize_text(value).upper()


def normalize_duplicate(value):
    text = normalize_text(value).upper()

    if text == "":
        return ""

    try:
        number = float(text)
        if number.is_integer():
            return str(int(number))
    except ValueError:
        pass

    return text


def normalize_ph(value):
    text = normalize_text(value)

    if text == "":
        return ""

    return text.replace(",", ".")


def parse_ph_number(value):
    text = normalize_ph(value)

    if text == "":
        return None

    if set(text) <= {"-"}:
        return None

    try:
        return float(text)
    except ValueError:
        return None


def extract_su_number(value):
    """
    Examples:
        SU1149-ILL-26  -> 1149
        'SU1149-ILL-26 -> 1149
        SU0079         -> 79
        SU0XXX-ILL-26  -> None
    """
    text = normalize_code(value)

    match = re.search(r"SU\s*0*(\d+)", text)

    if not match:
        return None

    return int(match.group(1))


def is_valid_ph_row(raw_code, raw_ph):
    code = normalize_code(raw_code)
    ph_number = parse_ph_number(raw_ph)
    su_number = extract_su_number(code)

    if code == "":
        return False, "empty code"

    if code.count("-") >= 3:
        return False, "too many dashes in code"

    if su_number is None:
        return False, "no valid SU number"

    if ph_number is None:
        return False, "no valid numeric pH"

    return True, ""


# ============================================================
# VERSION / SHEET RULES
# ============================================================

def detect_version_from_filename(path):
    name = path.name.upper()

    if "VER.03" in name or "VER 03" in name or "VER03" in name:
        return "Ver.03"

    if "VER.02" in name or "VER 02" in name or "VER02" in name:
        return "Ver.02"

    if "VER.01" in name or "VER 01" in name or "VER01" in name:
        return "Ver.01"

    return "Unknown"


def first_row_for_version(version):
    if version == "Ver.03":
        return 27

    if version in ("Ver.01", "Ver.02"):
        return 30

    return 30


def sheet_name_for_version(version):
    if version == "Ver.03":
        return "F-103"

    if version in ("Ver.01", "Ver.02"):
        return "F-41"

    return "F-41"


# ============================================================
# FILE DISCOVERY
# ============================================================

def get_excel_files(ph_folder, recursive=True, log=default_log):
    ph_folder = Path(ph_folder)

    if recursive:
        files = sorted(
            list(ph_folder.rglob("*.xlsx")) +
            list(ph_folder.rglob("*.xlsm")) +
            list(ph_folder.rglob("*.xls"))
        )
    else:
        files = sorted(
            list(ph_folder.glob("*.xlsx")) +
            list(ph_folder.glob("*.xlsm")) +
            list(ph_folder.glob("*.xls"))
        )

    selected = []

    for path in files:
        version = detect_version_from_filename(path)

        if version in ("Ver.01", "Ver.02", "Ver.03"):
            selected.append(path)

    log("")
    log("File discovery:")
    log(f"  Base folder:             {ph_folder}")
    log(f"  Recursive search:        {recursive}")
    log(f"  Excel files found:       {len(files)}")
    log(f"  Ver.01/02/03 selected:   {len(selected)}")

    return selected


# ============================================================
# CACHE HELPERS
# ============================================================

def file_cache_key(excel_file):
    excel_file = Path(excel_file)
    stat = excel_file.stat()

    raw = f"{excel_file.resolve()}|{stat.st_size}|{stat.st_mtime}".encode("utf-8")
    return hashlib.md5(raw).hexdigest()


def cached_xlsx_path(excel_file, cache_dir):
    cache_dir = Path(cache_dir)
    cache_dir.mkdir(parents=True, exist_ok=True)

    key = file_cache_key(excel_file)
    safe_stem = re.sub(r"[^A-Za-z0-9_.-]+", "_", Path(excel_file).stem)

    return cache_dir / f"{safe_stem}_{key}.xlsx"


# ============================================================
# LIBREOFFICE HELPERS
# ============================================================

def find_soffice():
    possible_paths = [
        "soffice",
        "libreoffice",
        "/Applications/LibreOffice.app/Contents/MacOS/soffice",
    ]

    for candidate in possible_paths:
        try:
            result = subprocess.run(
                [candidate, "--version"],
                stdout=subprocess.PIPE,
                stderr=subprocess.PIPE,
                text=True,
                timeout=20,
            )

            if result.returncode == 0:
                return candidate

        except Exception:
            pass

    return None


def convert_excel_with_libreoffice(excel_file, log):
    excel_file = Path(excel_file)

    soffice = find_soffice()

    if soffice is None:
        raise RuntimeError(
            "LibreOffice was not found. Install it with:\n"
            "  brew install --cask libreoffice"
        )

    temp_dir = Path(tempfile.mkdtemp(prefix="converted_excel_"))

    log("")
    log("Trying LibreOffice conversion:")
    log(f"  Original: {excel_file}")
    log(f"  Temp dir: {temp_dir}")

    cmd = [
        soffice,
        "--headless",
        "--convert-to",
        "xlsx",
        "--outdir",
        str(temp_dir),
        str(excel_file),
    ]

    result = subprocess.run(
        cmd,
        stdout=subprocess.PIPE,
        stderr=subprocess.PIPE,
        text=True,
        timeout=180,
    )

    if result.stdout.strip():
        log("LibreOffice stdout:")
        log(result.stdout.strip())

    if result.stderr.strip():
        log("LibreOffice stderr:")
        log(result.stderr.strip())

    if result.returncode != 0:
        raise RuntimeError(f"LibreOffice conversion failed for file: {excel_file}")

    converted_files = list(temp_dir.glob("*.xlsx"))

    if not converted_files:
        raise RuntimeError(
            f"LibreOffice did not create a converted .xlsx file for: {excel_file}"
        )

    return converted_files[0]


# ============================================================
# PASSWORD / DECRYPTION HELPERS
# ============================================================

def decrypt_excel_to_bytes(excel_file, password, log):
    if msoffcrypto is None:
        raise RuntimeError(
            "msoffcrypto-tool is not installed. Install it with:\n"
            "  pip install msoffcrypto-tool"
        )

    excel_file = Path(excel_file)

    log("")
    log("Trying password decryption:")
    log(f"  File: {excel_file}")
    log(f"  Password: {password!r}")

    decrypted = io.BytesIO()

    with open(excel_file, "rb") as f:
        office_file = msoffcrypto.OfficeFile(f)
        office_file.load_key(password=password)
        office_file.decrypt(decrypted)

    decrypted.seek(0)

    log("Password decryption succeeded.")
    return decrypted.getvalue()


def open_workbook_robust(
    excel_file,
    password,
    log,
    use_cache=True,
    cache_dir=".ph_excel_cache",
):
    """
    Fast opener.

    Tries:
      1. Cached clean xlsx
      2. openpyxl directly
      3. decrypt with password, save clean xlsx to cache
      4. LibreOffice conversion, save to cache
      5. LibreOffice conversion + password decrypt, save to cache

    Cache makes the second run much faster.
    """
    excel_file = Path(excel_file)

    cache_path = cached_xlsx_path(excel_file, cache_dir)

    # ------------------------------------------------------------
    # 1. Try cached clean file
    # ------------------------------------------------------------
    if use_cache and cache_path.exists():
        try:
            log("")
            log("Using cached recovered workbook:")
            log(f"  {cache_path}")

            wb = load_workbook(cache_path, data_only=True, read_only=True)
            return wb

        except Exception as e:
            log("")
            log("Cached file failed, deleting cache:")
            log(f"  {e}")

            try:
                cache_path.unlink()
            except Exception:
                pass

    # ------------------------------------------------------------
    # 2. Try openpyxl directly
    # ------------------------------------------------------------
    try:
        log("")
        log("Trying openpyxl directly...")
        wb = load_workbook(excel_file, data_only=True, read_only=True)
        log("Opened directly with openpyxl.")
        return wb

    except Exception as e_direct:
        log("")
        log("Direct openpyxl failed:")
        log(f"  {e_direct}")

    # ------------------------------------------------------------
    # 3. Try password decryption
    # ------------------------------------------------------------
    try:
        decrypted_bytes = decrypt_excel_to_bytes(excel_file, password, log)

        if use_cache:
            cache_path.write_bytes(decrypted_bytes)
            log("")
            log("Saved decrypted workbook to cache:")
            log(f"  {cache_path}")

            wb = load_workbook(cache_path, data_only=True, read_only=True)
        else:
            bio = io.BytesIO(decrypted_bytes)
            wb = load_workbook(bio, data_only=True, read_only=True)

        log("Opened after password decryption.")
        return wb

    except Exception as e_decrypt:
        log("")
        log("Password decryption path failed:")
        log(f"  {e_decrypt}")

    # ------------------------------------------------------------
    # 4. Try LibreOffice conversion
    # ------------------------------------------------------------
    converted_file = None

    try:
        converted_file = convert_excel_with_libreoffice(excel_file, log)

        if use_cache:
            cache_path.write_bytes(Path(converted_file).read_bytes())
            log("")
            log("Saved LibreOffice recovered workbook to cache:")
            log(f"  {cache_path}")

            wb = load_workbook(cache_path, data_only=True, read_only=True)
        else:
            wb = load_workbook(converted_file, data_only=True, read_only=True)

        log("Opened after LibreOffice conversion.")
        return wb

    except Exception as e_convert:
        log("")
        log("LibreOffice conversion path failed:")
        log(f"  {e_convert}")

    # ------------------------------------------------------------
    # 5. Try LibreOffice conversion + password decryption
    # ------------------------------------------------------------
    if converted_file is not None:
        try:
            decrypted_bytes = decrypt_excel_to_bytes(converted_file, password, log)

            if use_cache:
                cache_path.write_bytes(decrypted_bytes)
                log("")
                log("Saved LibreOffice + decrypted workbook to cache:")
                log(f"  {cache_path}")

                wb = load_workbook(cache_path, data_only=True, read_only=True)
            else:
                bio = io.BytesIO(decrypted_bytes)
                wb = load_workbook(bio, data_only=True, read_only=True)

            log("Opened after LibreOffice conversion + password decryption.")
            return wb

        except Exception as e_convert_decrypt:
            log("")
            log("LibreOffice + password path failed:")
            log(f"  {e_convert_decrypt}")

    raise RuntimeError(
        f"Could not open Excel file even after direct open, password '{password}', "
        f"and LibreOffice recovery: {excel_file}"
    )


# ============================================================
# DATABASE
# ============================================================

def create_database(db_file):
    conn = sqlite3.connect(db_file)
    cur = conn.cursor()

    cur.execute("DROP TABLE IF EXISTS ph_data")

    cur.execute(
        """
        CREATE TABLE ph_data (
            id INTEGER PRIMARY KEY AUTOINCREMENT,

            duplicate TEXT,
            code TEXT,
            ph REAL,

            su_number INTEGER,
            version TEXT,

            source_file TEXT,
            source_sheet TEXT,
            source_row INTEGER
        )
        """
    )

    cur.execute("CREATE INDEX idx_ph_code ON ph_data (code)")
    cur.execute("CREATE INDEX idx_ph_su_number ON ph_data (su_number)")
    cur.execute("CREATE INDEX idx_ph_version ON ph_data (version)")

    conn.commit()
    return conn


def insert_record(
    conn,
    duplicate,
    code,
    ph,
    su_number,
    version,
    source_file,
    source_sheet,
    source_row,
):
    cur = conn.cursor()

    cur.execute(
        """
        INSERT INTO ph_data (
            duplicate,
            code,
            ph,
            su_number,
            version,
            source_file,
            source_sheet,
            source_row
        )
        VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (
            duplicate,
            code,
            ph,
            su_number,
            version,
            source_file,
            source_sheet,
            source_row,
        ),
    )


# ============================================================
# FAST BLOCK READER
# ============================================================

def read_ph_blocks_from_sheet(
    ws,
    first_row,
    block_size,
    gap_rows,
    duplicate_col,
    code_col,
    ph_value_col,
    version,
    source_file_relative,
    sheet_name,
    best_records,
    conn,
    writer,
    keep_largest_ph_for_same_su,
    verbose,
    print_found,
    log,
):
    """
    Reads only real pH blocks.

    Example:
      Ver.03 first row 27:
        27-47
        50-70
        73-93
        ...

      Ver.01/02 first row 30:
        30-50
        53-73
        76-96
        ...
    """

    def vprint(message):
        if verbose:
            log(message)

    file_valid_rows = 0
    file_skipped_rows = 0
    total_inserted_now = 0

    max_empty_blocks_in_a_row = 3
    empty_blocks_in_a_row = 0

    block_start = first_row

    while block_start <= ws.max_row:
        block_end = block_start + block_size - 1

        # Quick check if this block contains real data
        valid_rows_in_block = 0

        for test_row in range(block_start, block_end + 1):
            raw_code_test = ws[f"{code_col}{test_row}"].value
            raw_ph_test = ws[f"{ph_value_col}{test_row}"].value

            valid_test, _ = is_valid_ph_row(raw_code_test, raw_ph_test)

            if valid_test:
                valid_rows_in_block += 1

        if valid_rows_in_block == 0:
            empty_blocks_in_a_row += 1

            if verbose:
                vprint("")
                vprint(f"Empty pH block: rows {block_start}-{block_end}")

            if empty_blocks_in_a_row >= max_empty_blocks_in_a_row:
                log("")
                log(f"Stopping sheet after {max_empty_blocks_in_a_row} empty pH blocks in a row.")
                break

            block_start = block_end + gap_rows + 1
            continue

        empty_blocks_in_a_row = 0

        if verbose:
            vprint("")
            vprint(f"Reading pH block rows {block_start}-{block_end}")
            vprint(f"Valid rows in block: {valid_rows_in_block}")

        for row in range(block_start, block_end + 1):
            duplicate_cell = f"{duplicate_col}{row}"
            code_cell = f"{code_col}{row}"
            ph_cell = f"{ph_value_col}{row}"

            raw_duplicate = ws[duplicate_cell].value
            raw_code = ws[code_cell].value
            raw_ph = ws[ph_cell].value

            duplicate = normalize_duplicate(raw_duplicate)
            code = normalize_code(raw_code)
            ph_number = parse_ph_number(raw_ph)
            su_number = extract_su_number(code)

            valid, reason = is_valid_ph_row(raw_code, raw_ph)

            if not valid:
                file_skipped_rows += 1

                if verbose:
                    vprint("")
                    vprint(f"Skipping row {row}: {reason}")
                    vprint(f"  {code_cell} code raw = {raw_code!r} -> {code!r}")
                    vprint(f"  {ph_cell} pH raw     = {raw_ph!r}")

                continue

            record = {
                "duplicate": duplicate,
                "code": code,
                "ph": ph_number,
                "su_number": su_number,
                "version": version,
                "source_file": source_file_relative,
                "source_sheet": sheet_name,
                "source_row": row,
            }

            file_valid_rows += 1

            if print_found:
                log(
                    f"FOUND: {code}  pH={ph_number}  "
                    f"version={version}  sheet={sheet_name}  row={row}"
                )

            if keep_largest_ph_for_same_su:
                old = best_records.get(su_number)

                if old is None or ph_number > old["ph"]:
                    best_records[su_number] = record

                    if verbose:
                        vprint("")
                        vprint(f"Selected/updated SU{su_number}:")
                        vprint(f"  code = {code}")
                        vprint(f"  pH   = {ph_number}")
                        vprint(f"  file = {source_file_relative}")
                        vprint(f"  row  = {row}")

            else:
                insert_record(
                    conn=conn,
                    duplicate=duplicate,
                    code=code,
                    ph=ph_number,
                    su_number=su_number,
                    version=version,
                    source_file=source_file_relative,
                    source_sheet=sheet_name,
                    source_row=row,
                )

                writer.writerow(
                    [
                        duplicate,
                        code,
                        ph_number,
                        su_number,
                        version,
                        source_file_relative,
                        sheet_name,
                        row,
                    ]
                )

                total_inserted_now += 1

        block_start = block_end + gap_rows + 1

    return file_valid_rows, file_skipped_rows, total_inserted_now


# ============================================================
# MAIN BUILDER
# ============================================================

def build_ph_database(
    ph_folder,
    database_file,
    csv_file,
    duplicate_col="B",
    code_col="C",
    ph_value_col="H",
    recursive=True,
    password="12",
    keep_largest_ph_for_same_su=True,
    verbose=True,
    print_found=False,
    use_cache=True,
    cache_dir=".ph_excel_cache",
    block_size=21,
    gap_rows=2,
    log_callback=None,
):
    log = log_callback if log_callback is not None else default_log

    ph_folder = Path(ph_folder)
    db_path = Path(database_file)
    csv_path = Path(csv_file)

    db_path.parent.mkdir(parents=True, exist_ok=True)
    csv_path.parent.mkdir(parents=True, exist_ok=True)

    log("")
    log("============================================================")
    log("BUILDING PH DATABASE FROM ALL EXCEL VERSIONS")
    log("============================================================")
    log(f"pH folder:         {ph_folder}")
    log(f"SQLite database:   {db_path}")
    log(f"CSV output:        {csv_path}")
    log(f"Cache enabled:     {use_cache}")
    log(f"Cache folder:      {cache_dir}")
    log(f"Print found rows:  {print_found}")
    log("")
    log("Version sheet / row rules:")
    log("  Ver.03 -> sheet F-103, starts C27 / H27")
    log("  Ver.02 -> sheet F-41,  starts C30 / H30")
    log("  Ver.01 -> sheet F-41,  starts C30 / H30")
    log("")
    log("Fast block reading:")
    log(f"  Block size: {block_size}")
    log(f"  Gap rows:   {gap_rows}")
    log("")
    log("Password handling:")
    log(f"  Password tried: {password!r}")
    log("============================================================")

    if not ph_folder.exists():
        raise FileNotFoundError(f"pH folder not found: {ph_folder}")

    excel_files = get_excel_files(
        ph_folder=ph_folder,
        recursive=recursive,
        log=log,
    )

    conn = create_database(db_path)

    csv_handle = open(csv_path, "w", newline="", encoding="utf-8-sig")
    writer = csv.writer(csv_handle)

    writer.writerow(
        [
            "duplicate",
            "code",
            "ph",
            "su_number",
            "version",
            "source_file",
            "source_sheet",
            "source_row",
        ]
    )

    total_inserted = 0
    total_skipped_rows = 0
    total_errors = 0
    total_files_opened = 0

    best_records = {}

    for file_index, excel_file in enumerate(excel_files, start=1):
        version = detect_version_from_filename(excel_file)
        first_row = first_row_for_version(version)
        sheet_name = sheet_name_for_version(version)

        try:
            source_file_relative = str(excel_file.relative_to(ph_folder))
        except ValueError:
            source_file_relative = str(excel_file)

        log("")
        log("============================================================")
        log(f"[{file_index}/{len(excel_files)}] ACCESSING EXCEL FILE")
        log("============================================================")
        log(f"File:    {excel_file}")
        log(f"Version: {version}")
        log(f"Sheet:   {sheet_name}")
        log(f"Start:   {code_col}{first_row} / {ph_value_col}{first_row}")

        try:
            wb = open_workbook_robust(
                excel_file=excel_file,
                password=password,
                log=log,
                use_cache=use_cache,
                cache_dir=cache_dir,
            )

            total_files_opened += 1

            if sheet_name not in wb.sheetnames:
                log("")
                log("SKIPPED FILE CONTENT:")
                log(f"  Expected sheet '{sheet_name}' for {version}, but it was not found.")
                log("Available sheets:")
                for sheet in wb.sheetnames:
                    log(f"  - {sheet}")

                wb.close()
                continue

            ws = wb[sheet_name]

            log("")
            log(f"Accessing sheet: {sheet_name}")
            log(f"Excel reported max row: {ws.max_row}")

            file_valid_rows, file_skipped_rows, inserted_now = read_ph_blocks_from_sheet(
                ws=ws,
                first_row=first_row,
                block_size=block_size,
                gap_rows=gap_rows,
                duplicate_col=duplicate_col,
                code_col=code_col,
                ph_value_col=ph_value_col,
                version=version,
                source_file_relative=source_file_relative,
                sheet_name=sheet_name,
                best_records=best_records,
                conn=conn,
                writer=writer,
                keep_largest_ph_for_same_su=keep_largest_ph_for_same_su,
                verbose=verbose,
                print_found=print_found,
                log=log,
            )

            total_skipped_rows += file_skipped_rows
            total_inserted += inserted_now

            wb.close()

            log("")
            log("------------------------------------------------------------")
            log(f"Finished file: {excel_file.name}")
            log(f"Valid pH rows found: {file_valid_rows}")
            log(f"Skipped rows:         {file_skipped_rows}")
            log("------------------------------------------------------------")

        except Exception as e:
            total_errors += 1

            log("")
            log("ERROR: Could not read this file even after recovery attempts:")
            log(f"  {excel_file}")
            log("Exception:")
            log(f"  {e}")

    if keep_largest_ph_for_same_su:
        log("")
        log("============================================================")
        log("WRITING BEST RECORDS TO DATABASE")
        log("============================================================")
        log("Rule:")
        log("  If same SU appears more than once, keep largest pH.")

        for su_number in sorted(best_records):
            record = best_records[su_number]

            insert_record(
                conn=conn,
                duplicate=record["duplicate"],
                code=record["code"],
                ph=record["ph"],
                su_number=record["su_number"],
                version=record["version"],
                source_file=record["source_file"],
                source_sheet=record["source_sheet"],
                source_row=record["source_row"],
            )

            writer.writerow(
                [
                    record["duplicate"],
                    record["code"],
                    record["ph"],
                    record["su_number"],
                    record["version"],
                    record["source_file"],
                    record["source_sheet"],
                    record["source_row"],
                ]
            )

            total_inserted += 1

    conn.commit()
    csv_handle.close()

    cur = conn.cursor()
    cur.execute("SELECT COUNT(*) FROM ph_data")
    db_count = cur.fetchone()[0]
    conn.close()

    log("")
    log("============================================================")
    log("DATABASE BUILD FINISHED")
    log("============================================================")
    log(f"Files selected:       {len(excel_files)}")
    log(f"Files opened:         {total_files_opened}")
    log(f"Total inserted rows:  {total_inserted}")
    log(f"Rows in SQLite DB:    {db_count}")
    log(f"Skipped rows:         {total_skipped_rows}")
    log(f"Files with errors:    {total_errors}")
    log("")
    log("Created:")
    log(f"  {db_path}")
    log(f"  {csv_path}")

    return {
        "files_selected": len(excel_files),
        "files_opened": total_files_opened,
        "total_inserted": total_inserted,
        "db_count": db_count,
        "skipped_rows": total_skipped_rows,
        "errors": total_errors,
        "database_file": str(db_path),
        "csv_file": str(csv_path),
    }


# ============================================================
# COMMAND LINE ARGUMENTS
# ============================================================

def parse_args():
    parser = argparse.ArgumentParser(
        description="Build pH SQLite database and CSV from Ver.01, Ver.02, and Ver.03 Excel files."
    )

    parser.add_argument(
        "--ph-folder",
        required=True,
        help="Folder containing the pH Excel files."
    )

    parser.add_argument(
        "--database-file",
        required=True,
        help="Output SQLite database path."
    )

    parser.add_argument(
        "--csv-file",
        required=True,
        help="Output CSV file path."
    )

    parser.add_argument(
        "--duplicate-col",
        default="B",
        help="Column for duplicate value. Default: B"
    )

    parser.add_argument(
        "--code-col",
        default="C",
        help="Column for SU code. Default: C"
    )

    parser.add_argument(
        "--ph-col",
        default="H",
        help="Column for pH value. Default: H"
    )

    parser.add_argument(
        "--password",
        default="12",
        help="Excel password to try. Default: 12"
    )

    parser.add_argument(
        "--no-recursive",
        action="store_true",
        help="Do not read subfolders such as Clientes."
    )

    parser.add_argument(
        "--keep-all-duplicates",
        action="store_true",
        help="Keep all duplicate SU records instead of keeping only the largest pH."
    )

    parser.add_argument(
        "--no-cache",
        action="store_true",
        help="Disable cached recovered/decrypted xlsx files."
    )

    parser.add_argument(
        "--cache-dir",
        default=".ph_excel_cache",
        help="Folder to store cached decrypted/recovered workbooks."
    )

    parser.add_argument(
        "--print-found",
        action="store_true",
        help="Print each valid SU code and pH while reading."
    )

    parser.add_argument(
        "--quiet",
        action="store_true",
        help="Reduce row-by-row logging."
    )

    return parser.parse_args()


def main():
    args = parse_args()

    build_ph_database(
        ph_folder=args.ph_folder,
        database_file=args.database_file,
        csv_file=args.csv_file,
        duplicate_col=args.duplicate_col,
        code_col=args.code_col,
        ph_value_col=args.ph_col,
        recursive=not args.no_recursive,
        password=args.password,
        keep_largest_ph_for_same_su=not args.keep_all_duplicates,
        verbose=not args.quiet,
        print_found=args.print_found,
        use_cache=not args.no_cache,
        cache_dir=args.cache_dir,
    )


if __name__ == "__main__":
    main()
