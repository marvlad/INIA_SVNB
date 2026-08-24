# build_ph_database.py
#
# Incremental pH database builder for INIA_SVNB
#
# Behaviour:
#
#   FIRST RUN WITH AN OLD DATABASE
#       - Detects that the old database does not have the
#         source-file tracking table.
#       - Rebuilds the pH data once.
#
#   SUBSEQUENT RUNS
#       - Checks every matching Excel file.
#       - If size + modification timestamp are unchanged:
#             SKIP
#       - If file is new:
#             PROCESS
#       - If file changed:
#             REPROCESS only that file
#       - If extraction settings changed:
#             REPROCESS affected files
#
#   CSV
#       - The complete CSV is exported from SQLite after syncing.
#       - This keeps compatibility with fill_ph_from_dat.py.
#
# Internal database tables:
#
#   ph_data
#       Actual pH measurements.
#
#   ph_source_files
#       Information about Excel files already processed.
#
# File identity:
#
#   source_key
#       Relative path inside the pH directory.
#
#       Example:
#           enero/F-103 Ver.03.xlsx
#
#       This is safer than using only the basename because two
#       different subfolders may contain files with the same name.
#
# CSV compatibility:
#
#   The CSV still contains:
#
#       duplicate
#       code
#       ph
#       su_number
#       source_file
#       source_sheet
#       source_row
#
#   Therefore fill_ph_from_dat.py does NOT need to change.


from pathlib import Path
from io import BytesIO

import argparse
import csv
import re
import sqlite3

import msoffcrypto

from openpyxl import load_workbook


# ============================================================
# PRINT / LOG HELPER
# ============================================================


def default_log(message):

    print(message)


# ============================================================
# EXCEL OPEN HELPER
# ============================================================


def open_excel_workbook(
    excel_file,
    password="12",
    data_only=True,
    read_only=True,
):

    """
    Open normal or password-protected Excel files.

    First try opening directly with openpyxl.

    If that fails, try decrypting with msoffcrypto
    using the supplied password.
    """

    try:

        return load_workbook(
            excel_file,
            data_only=data_only,
            read_only=read_only,
        )

    except Exception as normal_error:

        decrypted_file = BytesIO()

        try:

            with open(
                excel_file,
                "rb"
            ) as f:

                office_file = (
                    msoffcrypto.OfficeFile(f)
                )

                office_file.load_key(
                    password=password
                )

                office_file.decrypt(
                    decrypted_file
                )

            decrypted_file.seek(0)

            return load_workbook(
                decrypted_file,
                data_only=data_only,
                read_only=read_only,
            )

        except Exception as password_error:

            raise RuntimeError(
                "Could not open Excel file normally "
                f"or with password {password!r}. "
                f"Normal error: {normal_error}. "
                f"Password error: {password_error}"
            )


# ============================================================
# NORMALIZATION
# ============================================================


def normalize_text(value):

    """
    Clean Excel text.

    Handles things like:

        SU1149-ILL-26
        'SU1149-ILL-26
        extra spaces
        non-breaking spaces

    A leading Excel apostrophe is ignored.
    """

    if value is None:
        return ""

    text = str(value).strip()

    text = text.replace(
        "\xa0",
        " "
    )

    text = re.sub(
        r"\s+",
        " ",
        text
    )

    text = (
        text
        .lstrip("'")
        .strip()
    )

    return text


def normalize_code(value):

    """
    Normalize the SU code.
    """

    return normalize_text(
        value
    ).upper()


def normalize_duplicate(value):

    """
    Normalize duplicate value.

    Keeps values such as:

        D
        D2
        1
        2
        20

    Converts:

        1.0 -> 1
        2.0 -> 2
    """

    text = normalize_text(
        value
    ).upper()

    if text == "":
        return ""

    try:

        number = float(text)

        if number.is_integer():

            return str(
                int(number)
            )

    except ValueError:

        pass

    return text


def normalize_ph(value):

    """
    Normalize pH text.

    Converts:

        7,5 -> 7.5
    """

    text = normalize_text(
        value
    )

    if text == "":
        return ""

    text = text.replace(
        ",",
        "."
    )

    return text


def parse_ph_number(value):

    """
    Convert pH value to float.

    Valid examples:

        7
        7.1
        7,1
        8.25

    Invalid examples:

        ""
        "-"
        "---"
        "NO DATA"
        "abc"
        "S/D"
    """

    text = normalize_ph(
        value
    )

    if text == "":
        return None

    if set(text) <= {"-"}:
        return None

    try:

        return float(text)

    except ValueError:

        return None


def extract_su_number(value):

    """
    Extract numeric part of an SU code.

    Examples:

        SU1149-ILL-26 -> 1149
        'SU1149-ILL-26 -> 1149
        SU0079 -> 79
    """

    text = normalize_code(
        value
    )

    match = re.search(
        r"SU\s*0*(\d+)",
        text
    )

    if not match:
        return None

    return int(
        match.group(1)
    )


# ============================================================
# STOP RULE
# ============================================================


def should_stop_file(
    raw_code,
    raw_ph
):

    """
    Stop reading the current Excel file if ANY condition is true:

        1. Code is empty.

        2. Code contains at least three '-' characters.

        3. Code does not contain a valid SU number.

        4. pH does not contain a valid numeric value.
    """

    code = normalize_code(
        raw_code
    )

    ph_text = normalize_ph(
        raw_ph
    )

    su_number = extract_su_number(
        code
    )

    ph_number = parse_ph_number(
        raw_ph
    )

    if code == "":

        return (
            True,
            "Code column is empty"
        )

    if code.count("-") >= 3:

        return (
            True,
            "Code contains at least "
            f"3 '-' characters: {code!r}"
        )

    if su_number is None:

        return (
            True,
            "Code does not contain "
            f"a valid SU number: {code!r}"
        )

    if ph_number is None:

        return (
            True,
            "pH column does not contain "
            "a valid numeric value: "
            f"{ph_text!r}"
        )

    return False, ""


# ============================================================
# FILE FILTER
# ============================================================


def get_excel_files(
    ph_folder,
    file_name_filter,
    log,
):

    """
    Recursively find .xlsx and .xlsm files.

    Only files whose filename contains file_name_filter
    are returned.

    Comparison is case-insensitive.
    """

    all_excel_files = sorted(

        list(
            ph_folder.rglob(
                "*.xlsx"
            )
        )

        +

        list(
            ph_folder.rglob(
                "*.xlsm"
            )
        )
    )

    # --------------------------------------------------------
    # Ignore temporary Excel lock files.
    # --------------------------------------------------------

    all_excel_files = [

        path

        for path in all_excel_files

        if not path.name.startswith(
            "~$"
        )
    ]

    filter_upper = (
        file_name_filter
        .upper()
    )

    filtered_files = [

        path

        for path in all_excel_files

        if filter_upper
        in path.name.upper()
    ]

    log("")
    log("File filter:")

    log(
        "  Only reading files containing: "
        f"{file_name_filter}"
    )

    log(
        "  Total Excel files found: "
        f"{len(all_excel_files)}"
    )

    log(
        "  Selected files:          "
        f"{len(filtered_files)}"
    )

    log("")

    log(
        "Selected files:"
    )

    for path in filtered_files:

        try:

            relative = path.relative_to(
                ph_folder
            )

        except ValueError:

            relative = path

        log(
            f"  - {relative}"
        )

    return filtered_files


# ============================================================
# DATABASE SCHEMA
# ============================================================


def table_exists(
    conn,
    table_name,
):

    cur = conn.cursor()

    cur.execute(
        """
        SELECT 1
        FROM sqlite_master
        WHERE type = 'table'
        AND name = ?
        """,
        (
            table_name,
        ),
    )

    return (
        cur.fetchone()
        is not None
    )


def get_table_columns(
    conn,
    table_name,
):

    cur = conn.cursor()

    cur.execute(
        f"PRAGMA table_info({table_name})"
    )

    return {
        row[1]
        for row in cur.fetchall()
    }


def initialize_database(
    db_file
):

    """
    Create or migrate the SQLite database.

    Returns:

        conn,
        legacy_rebuild_required

    legacy_rebuild_required becomes True when an existing
    ph_data table is found but the source-file tracking table
    did not previously exist.

    In that situation we rebuild the pH data once because the
    old database cannot reliably tell us which rows came from
    which relative source file.
    """

    conn = sqlite3.connect(
        db_file
    )

    tracking_table_preexisting = (
        table_exists(
            conn,
            "ph_source_files",
        )
    )

    # --------------------------------------------------------
    # Main pH table.
    # --------------------------------------------------------

    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS ph_data (
            id INTEGER PRIMARY KEY AUTOINCREMENT,

            duplicate TEXT,
            code TEXT,
            ph REAL,

            su_number INTEGER,

            source_file TEXT,
            source_key TEXT,

            source_sheet TEXT,
            source_row INTEGER
        )
        """
    )

    # --------------------------------------------------------
    # Migration from the previous schema.
    #
    # Previous versions did not have source_key.
    # --------------------------------------------------------

    ph_columns = get_table_columns(
        conn,
        "ph_data"
    )

    if "source_key" not in ph_columns:

        conn.execute(
            """
            ALTER TABLE ph_data
            ADD COLUMN source_key TEXT
            """
        )

    # --------------------------------------------------------
    # Indexes.
    # --------------------------------------------------------

    conn.execute(
        """
        CREATE INDEX IF NOT EXISTS idx_ph_code
        ON ph_data (code)
        """
    )

    conn.execute(
        """
        CREATE INDEX IF NOT EXISTS idx_ph_su_number
        ON ph_data (su_number)
        """
    )

    conn.execute(
        """
        CREATE INDEX IF NOT EXISTS idx_ph_duplicate
        ON ph_data (duplicate)
        """
    )

    conn.execute(
        """
        CREATE INDEX IF NOT EXISTS idx_ph_source_file
        ON ph_data (source_file)
        """
    )

    conn.execute(
        """
        CREATE INDEX IF NOT EXISTS idx_ph_source_key
        ON ph_data (source_key)
        """
    )

    # --------------------------------------------------------
    # Source-file tracking table.
    # --------------------------------------------------------

    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS ph_source_files (

            source_key TEXT PRIMARY KEY,

            source_file TEXT NOT NULL,

            file_size INTEGER NOT NULL,

            mtime_ns INTEGER NOT NULL,

            parser_key TEXT NOT NULL,

            row_count INTEGER NOT NULL DEFAULT 0,

            processed_at TEXT NOT NULL
                DEFAULT CURRENT_TIMESTAMP
        )
        """
    )

    conn.commit()

    # --------------------------------------------------------
    # Detect old database requiring one-time migration.
    # --------------------------------------------------------

    legacy_rebuild_required = False

    if not tracking_table_preexisting:

        cur = conn.cursor()

        cur.execute(
            """
            SELECT COUNT(*)
            FROM ph_data
            """
        )

        existing_rows = (
            cur.fetchone()[0]
        )

        if existing_rows > 0:

            legacy_rebuild_required = True

    return (
        conn,
        legacy_rebuild_required,
    )


# ============================================================
# SOURCE FILE TRACKING
# ============================================================


def get_source_key(
    excel_file,
    ph_folder,
):

    """
    Return the file path relative to the pH root folder.

    Example:

        2026/enero/F-103 Ver.03.xlsx

    Path separators are normalized to '/'.
    """

    try:

        relative_path = (
            excel_file
            .resolve()
            .relative_to(
                ph_folder.resolve()
            )
        )

    except ValueError:

        relative_path = (
            Path(
                excel_file.name
            )
        )

    return (
        relative_path
        .as_posix()
    )


def get_file_signature(
    excel_file
):

    """
    Lightweight file signature.

    We deliberately avoid hashing the entire Excel file
    because these files may live on Google Drive/network
    storage.

    size + nanosecond modification time is normally sufficient
    to detect changes.
    """

    stat = excel_file.stat()

    return (
        stat.st_size,
        stat.st_mtime_ns,
    )


def make_parser_key(
    ph_sheet_name,
    ph_first_row,
    block_size,
    gap_rows,
    duplicate_col,
    code_col,
    ph_value_col,
):

    """
    Represent the extraction settings.

    If these settings change, previously processed files
    must be processed again.
    """

    parts = [

        f"sheet={ph_sheet_name}",

        f"first_row={int(ph_first_row)}",

        f"block_size={int(block_size)}",

        f"gap_rows={int(gap_rows)}",

        (
            "duplicate_col="
            f"{duplicate_col.upper()}"
        ),

        (
            "code_col="
            f"{code_col.upper()}"
        ),

        (
            "ph_value_col="
            f"{ph_value_col.upper()}"
        ),
    ]

    return "|".join(
        parts
    )


def file_needs_processing(
    conn,
    excel_file,
    source_key,
    parser_key,
):

    """
    Determine whether a source Excel file must be read.

    Returns:

        needs_processing,
        reason
    """

    file_size, mtime_ns = (
        get_file_signature(
            excel_file
        )
    )

    cur = conn.cursor()

    cur.execute(
        """
        SELECT
            file_size,
            mtime_ns,
            parser_key

        FROM ph_source_files

        WHERE source_key = ?
        """,
        (
            source_key,
        ),
    )

    existing = cur.fetchone()

    # --------------------------------------------------------
    # File has never been processed.
    # --------------------------------------------------------

    if existing is None:

        return (
            True,
            "NEW",
        )

    (
        old_file_size,
        old_mtime_ns,
        old_parser_key,
    ) = existing

    # --------------------------------------------------------
    # File changed.
    # --------------------------------------------------------

    if (
        old_file_size != file_size
        or old_mtime_ns != mtime_ns
    ):

        return (
            True,
            "CHANGED",
        )

    # --------------------------------------------------------
    # Extraction settings changed.
    # --------------------------------------------------------

    if (
        old_parser_key
        != parser_key
    ):

        return (
            True,
            "SETTINGS CHANGED",
        )

    return (
        False,
        "UNCHANGED",
    )


def mark_source_file_processed(
    conn,
    excel_file,
    source_key,
    parser_key,
    row_count,
):

    """
    Store/update information describing a successfully
    processed Excel source file.
    """

    file_size, mtime_ns = (
        get_file_signature(
            excel_file
        )
    )

    conn.execute(
        """
        INSERT INTO ph_source_files (

            source_key,
            source_file,

            file_size,
            mtime_ns,

            parser_key,

            row_count,

            processed_at
        )

        VALUES (
            ?, ?, ?, ?, ?, ?,
            CURRENT_TIMESTAMP
        )

        ON CONFLICT(source_key)

        DO UPDATE SET

            source_file =
                excluded.source_file,

            file_size =
                excluded.file_size,

            mtime_ns =
                excluded.mtime_ns,

            parser_key =
                excluded.parser_key,

            row_count =
                excluded.row_count,

            processed_at =
                CURRENT_TIMESTAMP
        """,
        (
            source_key,
            excel_file.name,

            file_size,
            mtime_ns,

            parser_key,

            int(row_count),
        ),
    )


# ============================================================
# DATABASE RECORD HELPERS
# ============================================================


def replace_source_records(
    conn,
    excel_file,
    source_key,
    parser_key,
    records,
):

    """
    Atomically replace all pH rows belonging to one source file.

    This is important for safety:

    - Excel is read first.
    - Records are collected in memory.
    - Only after extraction succeeds do we delete old rows.
    - New rows are inserted in the same transaction.

    Therefore an unreadable/corrupted Excel file does not destroy
    previously valid pH data.
    """

    with conn:

        conn.execute(
            """
            DELETE FROM ph_data
            WHERE source_key = ?
            """,
            (
                source_key,
            ),
        )

        conn.executemany(
            """
            INSERT INTO ph_data (

                duplicate,
                code,
                ph,

                su_number,

                source_file,
                source_key,

                source_sheet,
                source_row
            )

            VALUES (
                ?, ?, ?, ?, ?, ?, ?, ?
            )
            """,
            records,
        )

        mark_source_file_processed(
            conn=conn,
            excel_file=excel_file,
            source_key=source_key,
            parser_key=parser_key,
            row_count=len(records),
        )


def remove_stale_source_files(
    conn,
    selected_source_keys,
    log,
):

    """
    Remove database data belonging to source files that are no
    longer part of the selected Excel file set.

    Example:

        File was deleted from the pH folder.

    or:

        File no longer matches the configured filename filter.
    """

    cur = conn.cursor()

    cur.execute(
        """
        SELECT source_key
        FROM ph_source_files
        """
    )

    database_source_keys = {
        row[0]
        for row in cur.fetchall()
    }

    stale_keys = (
        database_source_keys
        - selected_source_keys
    )

    if not stale_keys:

        return 0

    log("")
    log(
        "Removing stale source files "
        "from database:"
    )

    with conn:

        for source_key in sorted(
            stale_keys
        ):

            log(
                f"  - {source_key}"
            )

            conn.execute(
                """
                DELETE FROM ph_data
                WHERE source_key = ?
                """,
                (
                    source_key,
                ),
            )

            conn.execute(
                """
                DELETE FROM ph_source_files
                WHERE source_key = ?
                """,
                (
                    source_key,
                ),
            )

    return len(
        stale_keys
    )


# ============================================================
# CSV EXPORT
# ============================================================


def export_database_to_csv(
    conn,
    csv_path,
):

    """
    Export the COMPLETE SQLite pH database to CSV.

    This keeps the existing CSV interface expected by
    fill_ph_from_dat.py.
    """

    columns = [

        "duplicate",
        "code",
        "ph",

        "su_number",

        "source_file",
        "source_sheet",
        "source_row",
    ]

    cur = conn.cursor()

    cur.execute(
        """
        SELECT

            duplicate,
            code,
            ph,

            su_number,

            source_file,
            source_sheet,
            source_row

        FROM ph_data

        ORDER BY

            source_key COLLATE NOCASE,

            source_row,

            id
        """
    )

    with open(
        csv_path,
        "w",
        newline="",
        encoding="utf-8-sig",
    ) as f:

        writer = csv.writer(
            f
        )

        writer.writerow(
            columns
        )

        for row in cur:

            writer.writerow(
                row
            )


# ============================================================
# EXCEL EXTRACTION
# ============================================================


def extract_records_from_excel(
    excel_file,
    source_key,
    ph_sheet_name,
    excel_password,
    ph_first_row,
    block_size,
    gap_rows,
    duplicate_col,
    code_col,
    ph_value_col,
    verbose,
    log,
):

    """
    Read one Excel workbook and return all pH records.

    No database changes are performed here.

    This allows us to safely replace the old database records
    only after the complete extraction succeeds.
    """

    records = []

    file_stopped = False

    stop_reason = ""

    wb = None

    try:

        log("")
        log(
            "Opening workbook..."
        )

        wb = open_excel_workbook(

            excel_file,

            password=excel_password,

            data_only=True,

            read_only=True,
        )

        log(
            "Workbook opened."
        )

        if verbose:

            log(
                "Available sheets:"
            )

            for sheet in wb.sheetnames:

                log(
                    f"  - {sheet}"
                )

        # ----------------------------------------------------
        # Required sheet.
        # ----------------------------------------------------

        if (
            ph_sheet_name
            not in wb.sheetnames
        ):

            raise RuntimeError(
                "Required sheet "
                f"{ph_sheet_name!r} "
                "was not found."
            )

        ws = wb[
            ph_sheet_name
        ]

        log("")
        log(
            f"Accessing sheet: "
            f"{ph_sheet_name}"
        )

        log(
            "Excel reported max row: "
            f"{ws.max_row}"
        )

        log(
            "Reading pH records until the "
            "first stop-rule row."
        )

        block_start = (
            ph_first_row
        )

        while True:

            block_end = (
                block_start
                + block_size
                - 1
            )

            if verbose:

                log("")
                log(
                    "------------------------------------------------------------"
                )

                log(
                    "Reading pH block: "
                    f"rows {block_start} "
                    f"to {block_end}"
                )

                log(
                    "------------------------------------------------------------"
                )

            stop_this_file = False

            for row in range(
                block_start,
                block_end + 1,
            ):

                duplicate_cell = (
                    f"{duplicate_col}{row}"
                )

                code_cell = (
                    f"{code_col}{row}"
                )

                ph_cell = (
                    f"{ph_value_col}{row}"
                )

                raw_duplicate = (
                    ws[
                        duplicate_cell
                    ].value
                )

                raw_code = (
                    ws[
                        code_cell
                    ].value
                )

                raw_ph = (
                    ws[
                        ph_cell
                    ].value
                )

                duplicate = (
                    normalize_duplicate(
                        raw_duplicate
                    )
                )

                code = (
                    normalize_code(
                        raw_code
                    )
                )

                ph_text = (
                    normalize_ph(
                        raw_ph
                    )
                )

                ph_number = (
                    parse_ph_number(
                        raw_ph
                    )
                )

                su_number = (
                    extract_su_number(
                        code
                    )
                )

                if verbose:

                    log("")
                    log(
                        f"Reading row {row}:"
                    )

                    log(
                        "  "
                        f"{duplicate_cell} "
                        "duplicate raw = "
                        f"{raw_duplicate!r} "
                        f"-> {duplicate!r}"
                    )

                    log(
                        "  "
                        f"{code_cell} "
                        "code raw      = "
                        f"{raw_code!r} "
                        f"-> {code!r}"
                    )

                    log(
                        "  "
                        f"{ph_cell} "
                        "pH raw        = "
                        f"{raw_ph!r} "
                        f"-> {ph_text!r}"
                    )

                    log(
                        "  pH numeric        = "
                        f"{ph_number}"
                    )

                    log(
                        "  Dash count        = "
                        f"{code.count('-')}"
                    )

                    log(
                        "  SU number         = "
                        f"{su_number}"
                    )

                stop, reason = (
                    should_stop_file(
                        raw_code,
                        raw_ph,
                    )
                )

                if stop:

                    log("")
                    log(
                        "STOP FILE:"
                    )

                    log(
                        f"  Reason: {reason}"
                    )

                    stop_this_file = True

                    file_stopped = True

                    stop_reason = reason

                    break

                # --------------------------------------------
                # Store record in memory.
                #
                # Database is NOT modified yet.
                # --------------------------------------------

                record = (

                    duplicate,
                    code,
                    ph_number,

                    su_number,

                    excel_file.name,
                    source_key,

                    ph_sheet_name,
                    row,
                )

                records.append(
                    record
                )

                if verbose:

                    log(
                        "  VALID RECORD:"
                    )

                    log(
                        "    duplicate = "
                        f"{duplicate}"
                    )

                    log(
                        "    code      = "
                        f"{code}"
                    )

                    log(
                        "    pH        = "
                        f"{ph_number}"
                    )

                    log(
                        "    row       = "
                        f"{row}"
                    )

            if stop_this_file:

                break

            # ------------------------------------------------
            # Move to next pH block.
            # ------------------------------------------------

            block_start = (
                block_end
                + gap_rows
                + 1
            )

            # ------------------------------------------------
            # Safety stop.
            # ------------------------------------------------

            if (
                block_start
                > ws.max_row
                + block_size
                + gap_rows
            ):

                log("")
                log(
                    "SAFETY STOP:"
                )

                log(
                    "  Reached beyond Excel's "
                    "reported maximum row."
                )

                stop_reason = (
                    "Safety stop beyond "
                    "Excel max row"
                )

                break

        return (
            records,
            file_stopped,
            stop_reason,
        )

    finally:

        if wb is not None:

            try:

                wb.close()

            except Exception:

                pass


# ============================================================
# MAIN BUILDER
# ============================================================


def build_ph_database(
    ph_folder,
    database_file,
    csv_file,
    file_name_filter="Ver.03",
    ph_sheet_name="F-103",
    excel_password="12",
    ph_first_row=27,
    block_size=21,
    gap_rows=2,
    duplicate_col="B",
    code_col="C",
    ph_value_col="H",
    verbose=True,
    log_callback=None,
):

    """
    Incrementally synchronize the pH SQLite database with
    the Excel files in ph_folder.

    Can be called from:

        1. command line
        2. Flask
        3. another Python program
    """

    log = (

        log_callback

        if log_callback is not None

        else default_log
    )

    ph_folder = Path(
        ph_folder
    )

    db_path = Path(
        database_file
    )

    csv_path = Path(
        csv_file
    )

    # --------------------------------------------------------
    # Make destination folders.
    # --------------------------------------------------------

    db_path.parent.mkdir(
        parents=True,
        exist_ok=True,
    )

    csv_path.parent.mkdir(
        parents=True,
        exist_ok=True,
    )

    # --------------------------------------------------------
    # Configuration report.
    # --------------------------------------------------------

    log("")
    log(
        "============================================================"
    )

    log(
        "SYNCING PH DATABASE FROM EXCEL FILES"
    )

    log(
        "============================================================"
    )

    log(
        "pH folder:"
    )

    log(
        f"  {ph_folder}"
    )

    log(
        "SQLite database:"
    )

    log(
        f"  {db_path}"
    )

    log(
        "CSV output:"
    )

    log(
        f"  {csv_path}"
    )

    log(
        "pH sheet:"
    )

    log(
        f"  {ph_sheet_name}"
    )

    log(
        "Filename filter:"
    )

    log(
        f"  {file_name_filter}"
    )

    log(
        "Columns:"
    )

    log(
        f"  {duplicate_col} = duplicate"
    )

    log(
        f"  {code_col} = code"
    )

    log(
        f"  {ph_value_col} = pH"
    )

    log(
        "Rows:"
    )

    log(
        f"  First row: {ph_first_row}"
    )

    log(
        f"  Block size: {block_size}"
    )

    log(
        f"  Gap rows: {gap_rows}"
    )

    log(
        "============================================================"
    )

    # --------------------------------------------------------
    # Validate pH folder.
    # --------------------------------------------------------

    if not ph_folder.exists():

        raise FileNotFoundError(
            "pH folder not found: "
            f"{ph_folder}"
        )

    if not ph_folder.is_dir():

        raise NotADirectoryError(
            "pH folder is not a directory: "
            f"{ph_folder}"
        )

    # --------------------------------------------------------
    # Find source Excel files.
    # --------------------------------------------------------

    excel_files = get_excel_files(
        ph_folder=ph_folder,
        file_name_filter=file_name_filter,
        log=log,
    )

    # --------------------------------------------------------
    # Open/create SQLite.
    # --------------------------------------------------------

    (
        conn,
        legacy_rebuild_required,
    ) = initialize_database(
        db_path
    )

    try:

        # ----------------------------------------------------
        # OLD DATABASE MIGRATION.
        #
        # The existing version of build_ph_database.py does
        # not maintain ph_source_files/source_key.
        #
        # Therefore the first execution of this version
        # performs one complete rebuild.
        # ----------------------------------------------------

        if legacy_rebuild_required:

            log("")
            log(
                "============================================================"
            )

            log(
                "OLD DATABASE FORMAT DETECTED"
            )

            log(
                "============================================================"
            )

            log(
                "The existing database was created "
                "before incremental file tracking."
            )

            log(
                "A one-time pH rebuild is required."
            )

            log(
                "After this run, unchanged Excel files "
                "will be skipped."
            )

            with conn:

                conn.execute(
                    """
                    DELETE FROM ph_data
                    """
                )

                conn.execute(
                    """
                    DELETE FROM ph_source_files
                    """
                )

        # ----------------------------------------------------
        # Parser signature.
        # ----------------------------------------------------

        parser_key = make_parser_key(

            ph_sheet_name=
                ph_sheet_name,

            ph_first_row=
                ph_first_row,

            block_size=
                block_size,

            gap_rows=
                gap_rows,

            duplicate_col=
                duplicate_col,

            code_col=
                code_col,

            ph_value_col=
                ph_value_col,
        )

        # ----------------------------------------------------
        # Calculate source keys.
        # ----------------------------------------------------

        source_info = []

        for excel_file in excel_files:

            source_key = (
                get_source_key(
                    excel_file,
                    ph_folder,
                )
            )

            source_info.append(
                (
                    excel_file,
                    source_key,
                )
            )

        selected_source_keys = {

            source_key

            for (
                excel_file,
                source_key,
            ) in source_info
        }

        # ----------------------------------------------------
        # Remove source files that disappeared.
        # ----------------------------------------------------

        stale_removed = (
            remove_stale_source_files(

                conn=conn,

                selected_source_keys=
                    selected_source_keys,

                log=log,
            )
        )

        # ----------------------------------------------------
        # Determine which Excel files actually need reading.
        # ----------------------------------------------------

        files_to_process = []

        unchanged_files = []

        log("")
        log(
            "============================================================"
        )

        log(
            "CHECKING SOURCE FILES"
        )

        log(
            "============================================================"
        )

        for (
            excel_file,
            source_key,
        ) in source_info:

            (
                needs_processing,
                reason,
            ) = file_needs_processing(

                conn=conn,

                excel_file=
                    excel_file,

                source_key=
                    source_key,

                parser_key=
                    parser_key,
            )

            if needs_processing:

                files_to_process.append(
                    (
                        excel_file,
                        source_key,
                        reason,
                    )
                )

                log(
                    f"PROCESS: {source_key} "
                    f"[{reason}]"
                )

            else:

                unchanged_files.append(
                    (
                        excel_file,
                        source_key,
                    )
                )

                log(
                    f"SKIP:    {source_key} "
                    "[UNCHANGED]"
                )

        log("")
        log(
            "Source file summary:"
        )

        log(
            "  Selected Excel files: "
            f"{len(excel_files)}"
        )

        log(
            "  Unchanged files skipped: "
            f"{len(unchanged_files)}"
        )

        log(
            "  New/changed files: "
            f"{len(files_to_process)}"
        )

        log(
            "  Stale files removed: "
            f"{stale_removed}"
        )

        # ----------------------------------------------------
        # Counters.
        # ----------------------------------------------------

        total_inserted = 0

        total_files_stopped = 0

        total_errors = 0

        processed_successfully = 0

        # ----------------------------------------------------
        # Process ONLY new or changed files.
        # ----------------------------------------------------

        for (
            file_index,
            file_data,
        ) in enumerate(
            files_to_process,
            start=1,
        ):

            (
                excel_file,
                source_key,
                update_reason,
            ) = file_data

            log("")
            log(
                "============================================================"
            )

            log(
                f"[{file_index}/"
                f"{len(files_to_process)}] "
                "PROCESSING EXCEL FILE"
            )

            log(
                "============================================================"
            )

            log(
                f"Reason: {update_reason}"
            )

            log(
                "File:"
            )

            log(
                f"  {excel_file}"
            )

            log(
                "Source key:"
            )

            log(
                f"  {source_key}"
            )

            try:

                (
                    records,
                    file_stopped,
                    stop_reason,
                ) = extract_records_from_excel(

                    excel_file=
                        excel_file,

                    source_key=
                        source_key,

                    ph_sheet_name=
                        ph_sheet_name,

                    excel_password=
                        excel_password,

                    ph_first_row=
                        ph_first_row,

                    block_size=
                        block_size,

                    gap_rows=
                        gap_rows,

                    duplicate_col=
                        duplicate_col,

                    code_col=
                        code_col,

                    ph_value_col=
                        ph_value_col,

                    verbose=
                        verbose,

                    log=
                        log,
                )

                # --------------------------------------------
                # IMPORTANT:
                #
                # Only now, after Excel extraction succeeded,
                # replace this file's database rows.
                # --------------------------------------------

                replace_source_records(

                    conn=conn,

                    excel_file=
                        excel_file,

                    source_key=
                        source_key,

                    parser_key=
                        parser_key,

                    records=
                        records,
                )

                file_inserted = len(
                    records
                )

                total_inserted += (
                    file_inserted
                )

                processed_successfully += 1

                if file_stopped:

                    total_files_stopped += 1

                log("")
                log(
                    "------------------------------------------------------------"
                )

                log(
                    "Finished file:"
                )

                log(
                    f"  {source_key}"
                )

                log(
                    "Rows inserted/replaced:"
                )

                log(
                    f"  {file_inserted}"
                )

                log(
                    "Stop reason:"
                )

                log(
                    f"  {stop_reason}"
                )

                log(
                    "------------------------------------------------------------"
                )

            except Exception as e:

                total_errors += 1

                log("")
                log(
                    "ERROR while reading file:"
                )

                log(
                    f"  {excel_file}"
                )

                log(
                    "Exception:"
                )

                log(
                    f"  {e}"
                )

                log("")
                log(
                    "IMPORTANT:"
                )

                log(
                    "  Previous valid SQLite rows for this "
                    "file were NOT deleted."
                )

                log(
                    "  The file will be tried again on the "
                    "next synchronization."
                )

        # ----------------------------------------------------
        # Export COMPLETE database to CSV.
        # ----------------------------------------------------

        log("")
        log(
            "============================================================"
        )

        log(
            "EXPORTING COMPLETE CSV FROM SQLITE"
        )

        log(
            "============================================================"
        )

        export_database_to_csv(
            conn=conn,
            csv_path=csv_path,
        )

        # ----------------------------------------------------
        # Final database totals.
        # ----------------------------------------------------

        cur = conn.cursor()

        cur.execute(
            """
            SELECT COUNT(*)
            FROM ph_data
            """
        )

        db_count = (
            cur.fetchone()[0]
        )

        cur.execute(
            """
            SELECT COUNT(*)
            FROM ph_source_files
            """
        )

        tracked_file_count = (
            cur.fetchone()[0]
        )

        # ----------------------------------------------------
        # Summary.
        # ----------------------------------------------------

        log("")
        log(
            "============================================================"
        )

        log(
            "PH DATABASE SYNC FINISHED"
        )

        log(
            "============================================================"
        )

        log(
            "Excel files selected: "
            f"{len(excel_files)}"
        )

        log(
            "Unchanged files skipped: "
            f"{len(unchanged_files)}"
        )

        log(
            "Files requiring processing: "
            f"{len(files_to_process)}"
        )

        log(
            "Files processed successfully: "
            f"{processed_successfully}"
        )

        log(
            "Stale source files removed: "
            f"{stale_removed}"
        )

        log(
            "Rows inserted/replaced this run: "
            f"{total_inserted}"
        )

        log(
            "Total rows in SQLite DB: "
            f"{db_count}"
        )

        log(
            "Tracked Excel files: "
            f"{tracked_file_count}"
        )

        log(
            "Files stopped by stop rule: "
            f"{total_files_stopped}"
        )

        log(
            "Files with errors: "
            f"{total_errors}"
        )

        log("")
        log(
            "SQLite:"
        )

        log(
            f"  {db_path}"
        )

        log(
            "CSV:"
        )

        log(
            f"  {csv_path}"
        )

        return {

            # Existing-style fields
            "total_inserted":
                total_inserted,

            "db_count":
                db_count,

            "files_stopped":
                total_files_stopped,

            "errors":
                total_errors,

            "database_file":
                str(db_path),

            "csv_file":
                str(csv_path),

            # New incremental-sync information
            "excel_files":
                len(excel_files),

            "unchanged_files":
                len(unchanged_files),

            "files_to_process":
                len(files_to_process),

            "processed_successfully":
                processed_successfully,

            "stale_files_removed":
                stale_removed,

            "tracked_files":
                tracked_file_count,
        }

    finally:

        conn.close()


# ============================================================
# COMMAND LINE ARGUMENTS
# ============================================================


def parse_args():

    parser = argparse.ArgumentParser(

        description=(
            "Incrementally synchronize the pH "
            "SQLite database and CSV from "
            "Excel Ver.03 files."
        )
    )

    parser.add_argument(

        "--ph-folder",

        required=True,

        help=(
            "Folder containing the pH "
            "Excel files."
        ),
    )

    parser.add_argument(

        "--database-file",

        required=True,

        help=(
            "SQLite database path."
        ),
    )

    parser.add_argument(

        "--csv-file",

        required=True,

        help=(
            "CSV output path."
        ),
    )

    parser.add_argument(

        "--file-name-filter",

        default="Ver.03",

        help=(
            "Only process Excel files whose "
            "names contain this text. "
            "Default: Ver.03"
        ),
    )

    parser.add_argument(

        "--sheet-name",

        default="F-103",

        help=(
            "Excel sheet name to read. "
            "Default: F-103"
        ),
    )

    parser.add_argument(

        "--excel-password",

        default="12",

        help=(
            "Password for protected Excel files. "
            "Default: 12"
        ),
    )

    parser.add_argument(

        "--first-row",

        type=int,

        default=27,

        help=(
            "First pH row to read. "
            "Default: 27"
        ),
    )

    parser.add_argument(

        "--block-size",

        type=int,

        default=21,

        help=(
            "Number of rows in each pH block. "
            "Default: 21"
        ),
    )

    parser.add_argument(

        "--gap-rows",

        type=int,

        default=2,

        help=(
            "Number of rows between blocks. "
            "Default: 2"
        ),
    )

    parser.add_argument(

        "--duplicate-col",

        default="B",

        help=(
            "Column containing duplicate value. "
            "Default: B"
        ),
    )

    parser.add_argument(

        "--code-col",

        default="C",

        help=(
            "Column containing SU code. "
            "Default: C"
        ),
    )

    parser.add_argument(

        "--ph-col",

        default="H",

        help=(
            "Column containing pH value. "
            "Default: H"
        ),
    )

    parser.add_argument(

        "--quiet",

        action="store_true",

        help=(
            "Reduce row-by-row logging."
        ),
    )

    return parser.parse_args()


# ============================================================
# MAIN
# ============================================================


def main():

    args = parse_args()

    build_ph_database(

        ph_folder=
            args.ph_folder,

        database_file=
            args.database_file,

        csv_file=
            args.csv_file,

        file_name_filter=
            args.file_name_filter,

        ph_sheet_name=
            args.sheet_name,

        excel_password=
            args.excel_password,

        ph_first_row=
            args.first_row,

        block_size=
            args.block_size,

        gap_rows=
            args.gap_rows,

        duplicate_col=
            args.duplicate_col,

        code_col=
            args.code_col,

        ph_value_col=
            args.ph_col,

        verbose=
            not args.quiet,
    )


if __name__ == "__main__":

    main()
