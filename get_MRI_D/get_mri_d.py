from pathlib import Path
import re
import io
import csv
import sqlite3
import argparse
from datetime import datetime

from openpyxl import load_workbook


try:
    import msoffcrypto
except ImportError:
    msoffcrypto = None


# ============================================================
# NORMALIZATION
# ============================================================

def normalize_text(value):
    if value is None:
        return ""

    text = str(value).strip()
    text = text.replace("\xa0", " ")
    text = re.sub(r"\s+", " ", text)
    text = text.lstrip("'").strip()

    return text


def normalize_code(value):
    return normalize_text(value).upper()


def normalize_tipo(value):
    return normalize_text(value).upper()


def parse_number(value):
    text = normalize_text(value)

    if text == "":
        return None

    text = text.replace(",", ".")

    if set(text) <= {"-"}:
        return None

    try:
        return float(text)
    except Exception:
        return None


def extract_result_value(value):
    """
    Keep whatever Excel has.

    Returns:
      resultado_text
      resultado_num

    Examples:
      5.23    -> ("5.23", 5.23)
      "5,23"  -> ("5,23", 5.23)
      "<0.01" -> ("<0.01", None)
      "ND"    -> ("ND", None)
    """
    text = normalize_text(value)

    if text == "":
        return "", None

    number = parse_number(value)
    return text, number


def is_mri_or_d(value):
    """
    Accept:
      MRI
      D
      D1
      D21
      D41
      etc.
    """
    tipo = normalize_tipo(value)

    if tipo == "MRI":
        return "MRI"

    if tipo == "D" or re.match(r"^D\d*$", tipo):
        return "D"

    return None


# ============================================================
# FILE / VERSION / METHOD
# ============================================================

def detect_version(path):
    name = Path(path).name.upper()

    if "VER.02" in name or "VER 02" in name or "VER02" in name:
        return "Ver.02"

    if "VER.04" in name or "VER 04" in name or "VER04" in name:
        return "Ver.04"

    if "VER.05" in name or "VER 05" in name or "VER05" in name:
        return "Ver.05"

    return "Unknown"


def detect_method(path):
    name = Path(path).name.upper()

    if "OLSEN" in name:
        return "OLSEN"

    if "BRAY" in name or "KURTZ" in name:
        return "BRAY Y KURTZ"

    return "UNKNOWN"


def settings_for_version(version):
    """
    Final fixed rules.

    Ver.02:
      data starts row = 30
      codigo = C
      tipo = E
      resultado = P
      fecha = E10
      password = 12

    Ver.04:
      data starts row = 31
      codigo = C
      tipo = F
      resultado = Q
      fecha = T12

    Ver.05:
      data starts row = 31
      codigo = C
      tipo = F
      resultado = Q
      fecha = T12
    """
    if version == "Ver.02":
        return {
            "password": "12",
            "date_cells": ["E10", "C10"],
            "header_row": 29,
            "start_row": 30,
            "codigo_col": "C",
            "tipo_col": "E",
            "resultado_col": "P",
        }

    if version == "Ver.04":
        return {
            "password": None,
            "date_cells": ["T12", "T13", "V4"],
            "header_row": 30,
            "start_row": 31,
            "codigo_col": "C",
            "tipo_col": "F",
            "resultado_col": "Q",
        }

    if version == "Ver.05":
        return {
            "password": None,
            "date_cells": ["T12", "T13", "V4"],
            "header_row": 30,
            "start_row": 31,
            "codigo_col": "C",
            "tipo_col": "F",
            "resultado_col": "Q",
        }

    return {
        "password": None,
        "date_cells": ["T12", "T13", "V4", "E10", "C10"],
        "header_row": 30,
        "start_row": 31,
        "codigo_col": "C",
        "tipo_col": "F",
        "resultado_col": "Q",
    }


# ============================================================
# DATE HELPERS
# ============================================================

def extract_date_from_filename(path):
    name = Path(path).name

    match = re.search(r"(\d{1,2})-(\d{1,2})-(\d{2,4})", name)

    if not match:
        return ""

    day = int(match.group(1))
    month = int(match.group(2))
    year = int(match.group(3))

    if year < 100:
        year += 2000

    try:
        return datetime(year, month, day).strftime("%Y-%m-%d")
    except Exception:
        return ""


def normalize_date(value, fallback=""):
    if value is None:
        return fallback

    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")

    text = normalize_text(value)

    if text == "":
        return fallback

    for fmt in (
        "%m/%d/%y",
        "%m/%d/%Y",
        "%d/%m/%y",
        "%d/%m/%Y",
        "%d-%m-%y",
        "%d-%m-%Y",
        "%Y-%m-%d",
    ):
        try:
            return datetime.strptime(text, fmt).strftime("%Y-%m-%d")
        except Exception:
            pass

    return text


def get_date_from_sheet(ws, date_cells, fallback):
    for cell in date_cells:
        try:
            value = ws[cell].value
        except Exception:
            continue

        date_value = normalize_date(value, fallback="")

        if date_value != "":
            return date_value, cell

    return fallback, "filename"


# ============================================================
# WORKBOOK OPENING
# ============================================================

def decrypt_excel_with_password(excel_file, password):
    if msoffcrypto is None:
        raise RuntimeError(
            "msoffcrypto-tool is not installed. Install it with:\n"
            "  pip install msoffcrypto-tool"
        )

    decrypted = io.BytesIO()

    with open(excel_file, "rb") as f:
        office_file = msoffcrypto.OfficeFile(f)
        office_file.load_key(password=password)
        office_file.decrypt(decrypted)

    decrypted.seek(0)

    return decrypted


def open_workbook_for_reading(excel_file, password=None, verbose=True):
    excel_file = Path(excel_file)
    keep_vba = excel_file.suffix.lower() == ".xlsm"

    try:
        if verbose:
            print("  Trying openpyxl directly...")

        wb = load_workbook(
            excel_file,
            data_only=True,
            keep_vba=keep_vba,
            read_only=False,
        )

        if verbose:
            print("  Opened directly.")

        return wb

    except Exception as direct_error:
        if verbose:
            print(f"  Direct open failed: {direct_error}")

        if password is None:
            raise

    if password is not None:
        if verbose:
            print(f"  Trying password decryption with password={password!r}...")

        decrypted = decrypt_excel_with_password(excel_file, password)

        wb = load_workbook(
            decrypted,
            data_only=True,
            keep_vba=keep_vba,
            read_only=False,
        )

        if verbose:
            print("  Opened after password decryption.")

        return wb

    raise RuntimeError(f"Could not open workbook: {excel_file}")


def get_sheet(wb, preferred_sheet="P_DIS"):
    if preferred_sheet in wb.sheetnames:
        return wb[preferred_sheet]

    return wb[wb.sheetnames[0]]


# ============================================================
# DATABASE
# ============================================================

def create_database(sqlite_path):
    sqlite_path = Path(sqlite_path)
    sqlite_path.parent.mkdir(parents=True, exist_ok=True)

    conn = sqlite3.connect(sqlite_path)
    cur = conn.cursor()

    cur.execute("DROP TABLE IF EXISTS colorimetric_results")

    cur.execute(
        """
        CREATE TABLE colorimetric_results (
            id INTEGER PRIMARY KEY AUTOINCREMENT,

            codigo_muestra TEXT,
            tipo TEXT,
            categoria TEXT,
            resultado TEXT,
            resultado_num REAL,

            metodo TEXT,
            version TEXT,
            fecha TEXT,
            date_source TEXT,

            source_file TEXT,
            source_sheet TEXT,
            source_row INTEGER,

            header_row INTEGER,
            start_row INTEGER,

            codigo_col TEXT,
            tipo_col TEXT,
            resultado_col TEXT,

            raw_tipo TEXT,
            raw_resultado TEXT,

            duplicated_from_source_row INTEGER
        )
        """
    )

    cur.execute("CREATE INDEX idx_color_code ON colorimetric_results (codigo_muestra)")
    cur.execute("CREATE INDEX idx_color_tipo ON colorimetric_results (tipo)")
    cur.execute("CREATE INDEX idx_color_categoria ON colorimetric_results (categoria)")
    cur.execute("CREATE INDEX idx_color_metodo ON colorimetric_results (metodo)")
    cur.execute("CREATE INDEX idx_color_version ON colorimetric_results (version)")
    cur.execute("CREATE INDEX idx_color_fecha ON colorimetric_results (fecha)")
    cur.execute("CREATE INDEX idx_color_source ON colorimetric_results (source_file, source_row)")

    conn.commit()

    return conn


def insert_record(conn, record):
    cur = conn.cursor()

    cur.execute(
        """
        INSERT INTO colorimetric_results (
            codigo_muestra,
            tipo,
            categoria,
            resultado,
            resultado_num,

            metodo,
            version,
            fecha,
            date_source,

            source_file,
            source_sheet,
            source_row,

            header_row,
            start_row,

            codigo_col,
            tipo_col,
            resultado_col,

            raw_tipo,
            raw_resultado,

            duplicated_from_source_row
        )
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (
            record["codigo_muestra"],
            record["tipo"],
            record["categoria"],
            record["resultado"],
            record["resultado_num"],

            record["metodo"],
            record["version"],
            record["fecha"],
            record["date_source"],

            record["source_file"],
            record["source_sheet"],
            record["source_row"],

            record["header_row"],
            record["start_row"],

            record["codigo_col"],
            record["tipo_col"],
            record["resultado_col"],

            record["raw_tipo"],
            record["raw_resultado"],

            record["duplicated_from_source_row"],
        ),
    )


# ============================================================
# RECORD MAKER
# ============================================================

def make_record_from_row(
    ws,
    row,
    categoria,
    codigo_col,
    tipo_col,
    resultado_col,
    metodo,
    version,
    fecha,
    date_source,
    source_file,
    header_row,
    start_row,
    duplicated_from_source_row=None,
):
    raw_codigo = ws[f"{codigo_col}{row}"].value
    raw_tipo = ws[f"{tipo_col}{row}"].value
    raw_resultado = ws[f"{resultado_col}{row}"].value

    codigo = normalize_code(raw_codigo)
    tipo = normalize_tipo(raw_tipo)
    resultado_text, resultado_num = extract_result_value(raw_resultado)

    if tipo == "" and categoria == "D_BELOW":
        tipo = "D_BELOW"

    if codigo == "":
        return None, "empty codigo"

    if resultado_text == "":
        return None, f"empty resultado at {resultado_col}{row}"

    record = {
        "codigo_muestra": codigo,
        "tipo": tipo,
        "categoria": categoria,
        "resultado": resultado_text,
        "resultado_num": resultado_num,

        "metodo": metodo,
        "version": version,
        "fecha": fecha,
        "date_source": date_source,

        "source_file": source_file,
        "source_sheet": ws.title,
        "source_row": row,

        "header_row": header_row,
        "start_row": start_row,

        "codigo_col": codigo_col,
        "tipo_col": tipo_col,
        "resultado_col": resultado_col,

        "raw_tipo": normalize_text(raw_tipo),
        "raw_resultado": normalize_text(raw_resultado),

        "duplicated_from_source_row": duplicated_from_source_row,
    }

    return record, ""


# ============================================================
# DEBUG HELPER
# ============================================================

def print_debug_rows(ws, start_row, codigo_col, tipo_col, resultado_col, n_rows=25):
    print("")
    print("DEBUG FIRST ROWS READ")
    print("------------------------------------------------------------")

    end_row = min(ws.max_row, start_row + n_rows - 1)

    for row in range(start_row, end_row + 1):
        print(
            f"DEBUG row={row} "
            f"{codigo_col}={ws[f'{codigo_col}{row}'].value!r} "
            f"{tipo_col}={ws[f'{tipo_col}{row}'].value!r} "
            f"{resultado_col}={ws[f'{resultado_col}{row}'].value!r}"
        )

    print("------------------------------------------------------------")


# ============================================================
# EXTRACTION
# ============================================================

def extract_records_from_workbook(
    excel_file,
    input_dir,
    sheet_name="P_DIS",
    verbose=True,
    debug_rows=False,
):
    excel_file = Path(excel_file)
    input_dir = Path(input_dir)

    version = detect_version(excel_file)
    metodo = detect_method(excel_file)
    settings = settings_for_version(version)

    password = settings["password"]
    date_cells = settings["date_cells"]
    header_row = settings["header_row"]
    start_row = settings["start_row"]
    codigo_col = settings["codigo_col"]
    tipo_col = settings["tipo_col"]
    resultado_col = settings["resultado_col"]

    print("")
    print("============================================================")
    print("EXTRACTING FILE")
    print("============================================================")
    print(f"File:           {excel_file}")
    print(f"Version:        {version}")
    print(f"Method:         {metodo}")
    print(f"Sheet:          {sheet_name}")
    print(f"Header row:     {header_row}")
    print(f"Data starts:    {start_row}")
    print(f"Codigo column:  {codigo_col}")
    print(f"Tipo column:    {tipo_col}")
    print(f"Result column:  {resultado_col}")
    print(f"Date cells:     {date_cells}")
    print(f"Password:       {password!r}")

    wb = open_workbook_for_reading(
        excel_file=excel_file,
        password=password,
        verbose=verbose,
    )

    ws = get_sheet(wb, preferred_sheet=sheet_name)

    fallback_date = extract_date_from_filename(excel_file)
    fecha, date_source = get_date_from_sheet(
        ws=ws,
        date_cells=date_cells,
        fallback=fallback_date,
    )

    try:
        source_file = str(excel_file.relative_to(input_dir))
    except Exception:
        source_file = str(excel_file)

    print(f"Date found:     {fecha}")
    print(f"Date source:    {date_source}")
    print(f"Using sheet:    {ws.title}")
    print(f"Max row:        {ws.max_row}")
    print(f"Max column:     {ws.max_column}")

    if debug_rows:
        print_debug_rows(
            ws=ws,
            start_row=start_row,
            codigo_col=codigo_col,
            tipo_col=tipo_col,
            resultado_col=resultado_col,
            n_rows=30,
        )

    records = []

    extracted_mri = 0
    extracted_d = 0
    extracted_d_below = 0
    skipped = 0

    for row in range(start_row, ws.max_row + 1):
        raw_tipo = ws[f"{tipo_col}{row}"].value
        tipo = normalize_tipo(raw_tipo)
        categoria_found = is_mri_or_d(tipo)

        if categoria_found not in ("MRI", "D"):
            continue

        raw_resultado = ws[f"{resultado_col}{row}"].value
        resultado_text, resultado_num = extract_result_value(raw_resultado)

        if resultado_text == "":
            skipped += 1
            if verbose:
                print(
                    f"  SKIP {categoria_found}: row={row} "
                    f"tipo={tipo!r} empty result at {resultado_col}{row}"
                )
            continue

        record, reason = make_record_from_row(
            ws=ws,
            row=row,
            categoria=categoria_found,
            codigo_col=codigo_col,
            tipo_col=tipo_col,
            resultado_col=resultado_col,
            metodo=metodo,
            version=version,
            fecha=fecha,
            date_source=date_source,
            source_file=source_file,
            header_row=header_row,
            start_row=start_row,
            duplicated_from_source_row=None,
        )

        if record is None:
            skipped += 1
            if verbose:
                print(f"  SKIP {categoria_found}: row={row}: {reason}")
            continue

        records.append(record)

        if categoria_found == "MRI":
            extracted_mri += 1
        elif categoria_found == "D":
            extracted_d += 1

        if verbose:
            print(
                f"  FOUND {categoria_found}: row={row} "
                f"codigo={record['codigo_muestra']} "
                f"tipo={tipo} "
                f"resultado={resultado_text} "
                f"resultado_num={resultado_num}"
            )

        # Extract row immediately below D
        if categoria_found == "D":
            below_row = row + 1

            if below_row <= ws.max_row:
                below_raw_resultado = ws[f"{resultado_col}{below_row}"].value
                below_resultado_text, below_resultado_num = extract_result_value(
                    below_raw_resultado
                )

                if below_resultado_text == "":
                    skipped += 1
                    if verbose:
                        print(
                            f"  SKIP D_BELOW: row={below_row} "
                            f"empty result at {resultado_col}{below_row}"
                        )
                else:
                    below_record, below_reason = make_record_from_row(
                        ws=ws,
                        row=below_row,
                        categoria="D_BELOW",
                        codigo_col=codigo_col,
                        tipo_col=tipo_col,
                        resultado_col=resultado_col,
                        metodo=metodo,
                        version=version,
                        fecha=fecha,
                        date_source=date_source,
                        source_file=source_file,
                        header_row=header_row,
                        start_row=start_row,
                        duplicated_from_source_row=row,
                    )

                    if below_record is None:
                        skipped += 1
                        if verbose:
                            print(f"  SKIP D_BELOW: row={below_row}: {below_reason}")
                    else:
                        records.append(below_record)
                        extracted_d_below += 1

                        if verbose:
                            print(
                                f"  FOUND D_BELOW: row={below_row} "
                                f"codigo={below_record['codigo_muestra']} "
                                f"resultado={below_record['resultado']} "
                                f"resultado_num={below_record['resultado_num']} "
                                f"duplicated_from_row={row}"
                            )

    wb.close()

    print("")
    print("Finished file:")
    print(f"  Extracted MRI:      {extracted_mri}")
    print(f"  Extracted D:        {extracted_d}")
    print(f"  Extracted D_BELOW:  {extracted_d_below}")
    print(f"  Skipped:            {skipped}")
    print(f"  Total records:      {len(records)}")

    return records


# ============================================================
# FILE DISCOVERY
# ============================================================

def get_excel_files(input_dir, min_number=None, max_number=None):
    input_dir = Path(input_dir)

    files = sorted(
        list(input_dir.glob("*.xlsx")) +
        list(input_dir.glob("*.xlsm"))
    )

    selected = []

    print("")
    print("Checking Excel files in:")
    print(f"  {input_dir}")
    print(f"Total Excel files found: {len(files)}")

    for path in files:
        if path.name.startswith("~"):
            print(f"Skipping temporary/hidden file: {path.name}")
            continue

        name_upper = path.name.upper()

        # Only process files that have SU in the filename.
        if "SU" not in name_upper:
            print(f"Skipping file without SU in name: {path.name}")
            continue

        version = detect_version(path)

        print(f"FOUND FILE: version={version} file={path.name}")

        if version not in ("Ver.02", "Ver.04", "Ver.05"):
            continue

        if min_number is not None or max_number is not None:
            match = re.match(r"^(\d+)", path.name.strip())

            if not match:
                continue

            number = int(match.group(1))

            if min_number is not None and number < min_number:
                continue

            if max_number is not None and number > max_number:
                continue

        selected.append(path)

    print("")
    print("============================================================")
    print("INPUT FILES")
    print("============================================================")
    print(f"Input dir:      {input_dir}")
    print(f"Selected files: {len(selected)}")

    for path in selected:
        print(f"  - {path.name}")

    return selected


# ============================================================
# CSV EXPORT
# ============================================================

def export_csv(records, csv_path):
    csv_path = Path(csv_path)
    csv_path.parent.mkdir(parents=True, exist_ok=True)

    fieldnames = [
        "codigo_muestra",
        "tipo",
        "categoria",
        "resultado",
        "resultado_num",
        "metodo",
        "version",
        "fecha",
        "date_source",
        "source_file",
        "source_sheet",
        "source_row",
        "header_row",
        "start_row",
        "codigo_col",
        "tipo_col",
        "resultado_col",
        "raw_tipo",
        "raw_resultado",
        "duplicated_from_source_row",
    ]

    with open(csv_path, "w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()

        for record in records:
            writer.writerow(record)


# ============================================================
# MAIN
# ============================================================

def build_colorimetric_database(
    input_dir,
    sqlite_path,
    csv_path=None,
    sheet_name="P_DIS",
    min_number=None,
    max_number=None,
    verbose=True,
    debug_rows=False,
):
    input_dir = Path(input_dir)
    sqlite_path = Path(sqlite_path)

    print("")
    print("============================================================")
    print("BUILD COLORIMETRIC MRI/D/D_BELOW DATABASE")
    print("============================================================")
    print(f"Input dir:  {input_dir}")
    print(f"SQLite DB:  {sqlite_path}")
    print(f"CSV output: {csv_path}")
    print("")
    print("Rules:")
    print("  Only files with SU in filename are processed")
    print("  Ver.02: data starts row 30, codigo=C, tipo=E, resultado=P")
    print("  Ver.04: data starts row 31, codigo=C, tipo=F, resultado=Q")
    print("  Ver.05: data starts row 31, codigo=C, tipo=F, resultado=Q")
    print("  resultado stored as text")
    print("  resultado_num stored only if numeric conversion works")
    print("  Ver.02 uses password 12")
    print("  Extract MRI")
    print("  Extract D, D1, D21, etc.")
    print("  Extract row immediately below D as D_BELOW")
    print("============================================================")

    if not input_dir.exists():
        raise FileNotFoundError(f"Input folder not found: {input_dir}")

    excel_files = get_excel_files(
        input_dir=input_dir,
        min_number=min_number,
        max_number=max_number,
    )

    conn = create_database(sqlite_path)

    all_records = []
    successful_files = 0
    failed_files = 0

    for index, excel_file in enumerate(excel_files, start=1):
        print("")
        print("############################################################")
        print(f"PROCESSING {index}/{len(excel_files)}")
        print("############################################################")

        try:
            records = extract_records_from_workbook(
                excel_file=excel_file,
                input_dir=input_dir,
                sheet_name=sheet_name,
                verbose=verbose,
                debug_rows=debug_rows,
            )

            for record in records:
                insert_record(conn, record)

            all_records.extend(records)
            successful_files += 1

        except Exception as e:
            failed_files += 1
            print("")
            print("ERROR PROCESSING FILE:")
            print(f"  {excel_file}")
            print(f"  {e}")

    conn.commit()
    conn.close()

    if csv_path is not None:
        export_csv(all_records, csv_path)

    print("")
    print("============================================================")
    print("DATABASE BUILD FINISHED")
    print("============================================================")
    print(f"Files selected:     {len(excel_files)}")
    print(f"Successful files:   {successful_files}")
    print(f"Failed files:       {failed_files}")
    print(f"Total records:      {len(all_records)}")
    print(f"SQLite created:     {sqlite_path}")

    if csv_path is not None:
        print(f"CSV created:        {csv_path}")


# ============================================================
# ARGUMENTS
# ============================================================

def parse_args():
    parser = argparse.ArgumentParser(
        description="Extract MRI, D, and D_BELOW rows from F-82 colorimetric Excel files into SQLite."
    )

    parser.add_argument(
        "--input-dir",
        default="/Users/mascenci/Downloads/Clientes",
        help="Folder containing colorimetric Excel files."
    )

    parser.add_argument(
        "--sqlite",
        default="database_colorimetric/colorimetric_mri_d.sqlite",
        help="Output SQLite database path."
    )

    parser.add_argument(
        "--csv",
        default="database_colorimetric/colorimetric_mri_d.csv",
        help="Optional output CSV path."
    )

    parser.add_argument(
        "--sheet-name",
        default="P_DIS",
        help="Sheet name to read. Default: P_DIS."
    )

    parser.add_argument(
        "--min-number",
        type=int,
        default=None,
        help="Optional first file number to process."
    )

    parser.add_argument(
        "--max-number",
        type=int,
        default=None,
        help="Optional last file number to process."
    )

    parser.add_argument(
        "--quiet",
        action="store_true",
        help="Reduce row-level logging."
    )

    parser.add_argument(
        "--debug-rows",
        action="store_true",
        help="Print first rows actually read for each file."
    )

    return parser.parse_args()


def main():
    args = parse_args()

    build_colorimetric_database(
        input_dir=args.input_dir,
        sqlite_path=args.sqlite,
        csv_path=args.csv,
        sheet_name=args.sheet_name,
        min_number=args.min_number,
        max_number=args.max_number,
        verbose=not args.quiet,
        debug_rows=args.debug_rows,
    )


if __name__ == "__main__":
    main()
