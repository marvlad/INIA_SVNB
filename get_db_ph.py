# build_ph_database_from_excel_files_Ver03_only.py

from pathlib import Path
from io import BytesIO
import re
import csv
import sqlite3

import msoffcrypto
from openpyxl import load_workbook


# ============================================================
# CONFIGURATION
# ============================================================

PH_FOLDER = r"G:\Mi unidad\LABSAF ILLPA\1. Documentos Internos\7.5 Registros Tecnicos\2026\SUELOS\1.pH"

DATABASE_FILE = r"G:\Mi unidad\LABSAF ILLPA\ph_database_Ver03.sqlite"

CSV_FILE = r"G:\Mi unidad\LABSAF ILLPA\ph_database_Ver03.csv"

# Only files containing this text in the filename will be processed.
FILE_NAME_FILTER = "Ver.03"

# pH Excel tab
PH_SHEET_NAME = "F-103"

# Password used only if the Excel file needs one.
EXCEL_PASSWORD = "12"

# pH file structure
PH_FIRST_ROW = 27

# Read blocks:
#   27 to 47
#   50 to 70
#   73 to 93
#   96 to 116
#   ...
BLOCK_SIZE = 21
GAP_ROWS = 2

PH_DUPLICATE_COL = "B"
PH_CODE_COL = "C"
PH_VALUE_COL = "H"

VERBOSE = True


# ============================================================
# PRINT HELPER
# ============================================================

def vprint(message):
    if VERBOSE:
        print(message)


# ============================================================
# EXCEL OPEN HELPER
# ============================================================

def open_excel_workbook(excel_file, password=EXCEL_PASSWORD):
    """
    Open an Excel workbook.

    First tries normally.
    If the file is password-protected/encrypted, it tries again using password="12".
    """
    try:
        return load_workbook(excel_file, data_only=True, read_only=True)

    except Exception as first_error:
        print("")
        print("Normal open failed. Trying with password...")

        try:
            decrypted = BytesIO()

            with open(excel_file, "rb") as file_handle:
                office_file = msoffcrypto.OfficeFile(file_handle)
                office_file.load_key(password=password)
                office_file.decrypt(decrypted)

            decrypted.seek(0)

            return load_workbook(decrypted, data_only=True, read_only=True)

        except Exception as second_error:
            raise RuntimeError(
                f"Could not open workbook normally or with password '{password}'. "
                f"Normal error: {first_error}. "
                f"Password error: {second_error}"
            )


# ============================================================
# NORMALIZATION
# ============================================================

def normalize_text(value):
    """
    Clean Excel text.

    Handles:
        SU1149-ILL-26
        'SU1149-ILL-26
        spaces
        non-breaking spaces

    The apostrophe in front is ignored.
    """
    if value is None:
        return ""

    text = str(value).strip()
    text = text.replace("\xa0", " ")
    text = re.sub(r"\s+", " ", text)

    # Ignore leading Excel apostrophe:
    #   'SU1149-ILL-26 -> SU1149-ILL-26
    text = text.lstrip("'").strip()

    return text


def normalize_code(value):
    """
    Normalize SU code from column C.
    """
    return normalize_text(value).upper()


def normalize_duplicate(value):
    """
    Normalize duplicate value from column B.

    Keeps:
        D
        D2
        1
        2
        20

    Converts:
        1.0 -> 1
        2.0 -> 2
    """
    text = normalize_text(value).upper()

    if text == "":
        return ""

    try:
        number = float(text)
        if number.is_integer():
            return str(int(number))
    except ValueError:
        pass

    return text


def normalize_ph(value):
    """
    Normalize pH value from column H.

    Converts comma decimal to dot:
        7,5 -> 7.5

    This only cleans the text.
    It does not decide if the value is valid.
    """
    text = normalize_text(value)

    if text == "":
        return ""

    text = text.replace(",", ".")

    return text


def parse_ph_number(value):
    """
    Convert pH value to float.

    Valid:
        7
        7.1
        7,1
        8.25

    Invalid:
        ""
        "-"
        "---"
        "NO DATA"
        "abc"
        "S/D"
    """
    text = normalize_ph(value)

    if text == "":
        return None

    # If it contains only dashes, it is not real data.
    # Examples:
    #   -
    #   --
    #   ---
    if set(text) <= {"-"}:
        return None

    try:
        return float(text)
    except ValueError:
        return None


def extract_su_number(value):
    """
    Extract numeric SU number.

    Examples:
        SU1149-ILL-26   -> 1149
        'SU1149-ILL-26  -> 1149
        SU0079          -> 79
    """
    text = normalize_code(value)

    match = re.search(r"SU\s*0*(\d+)", text)

    if not match:
        return None

    return int(match.group(1))


def should_stop_file(raw_code, raw_ph):
    """
    Decide when the useful data in the current pH Excel file ended.

    Stop if ANY of these is true:

      1. Column C has no code.
      2. Column C has at least 3 '-' characters.
      3. Column C does not contain a valid SU number.
      4. Column H has no meaningful numeric pH value.

    This is OR logic.
    If one condition is true, the script stops that file
    and moves to the next Excel file.
    """
    code = normalize_code(raw_code)
    ph_text = normalize_ph(raw_ph)

    su_number = extract_su_number(code)
    ph_number = parse_ph_number(raw_ph)

    if code == "":
        return True, "Column C is empty"

    if code.count("-") >= 3:
        return True, f"Column C has at least 3 '-' characters: {code!r}"

    if su_number is None:
        return True, f"Column C does not contain a valid SU code: {code!r}"

    if ph_number is None:
        return True, f"Column H does not contain a valid numeric pH value: {ph_text!r}"

    return False, ""


# ============================================================
# FILE FILTER
# ============================================================

def get_ver03_excel_files(ph_folder):
    """
    Return only .xlsx/.xlsm files whose name contains Ver.03.
    Case-insensitive.
    """
    all_excel_files = sorted(
        list(ph_folder.glob("*.xlsx")) +
        list(ph_folder.glob("*.xlsm"))
    )

    filtered_files = [
        path for path in all_excel_files
        if FILE_NAME_FILTER.upper() in path.name.upper()
    ]

    print("")
    print("File filter:")
    print(f"  Only reading files containing: {FILE_NAME_FILTER}")
    print(f"  Total Excel files found: {len(all_excel_files)}")
    print(f"  Ver.03 files selected:   {len(filtered_files)}")

    print("")
    print("Selected files:")
    for path in filtered_files:
        print(f"  - {path.name}")

    return filtered_files


# ============================================================
# DATABASE
# ============================================================

def create_database(db_file):
    conn = sqlite3.connect(db_file)
    cur = conn.cursor()

    cur.execute("DROP TABLE IF EXISTS ph_data")

    cur.execute(
        """
        CREATE TABLE ph_data (
            id INTEGER PRIMARY KEY AUTOINCREMENT,

            duplicate TEXT,
            code TEXT,
            ph REAL,

            su_number INTEGER,

            source_file TEXT,
            source_sheet TEXT,
            source_row INTEGER
        )
        """
    )

    cur.execute("CREATE INDEX idx_ph_code ON ph_data (code)")
    cur.execute("CREATE INDEX idx_ph_su_number ON ph_data (su_number)")
    cur.execute("CREATE INDEX idx_ph_duplicate ON ph_data (duplicate)")

    conn.commit()

    return conn


def insert_record(
    conn,
    duplicate,
    code,
    ph,
    su_number,
    source_file,
    source_sheet,
    source_row,
):
    cur = conn.cursor()

    cur.execute(
        """
        INSERT INTO ph_data (
            duplicate,
            code,
            ph,
            su_number,
            source_file,
            source_sheet,
            source_row
        )
        VALUES (?, ?, ?, ?, ?, ?, ?)
        """,
        (
            duplicate,
            code,
            ph,
            su_number,
            source_file,
            source_sheet,
            source_row,
        ),
    )


# ============================================================
# MAIN BUILDER
# ============================================================

def build_ph_database():
    ph_folder = Path(PH_FOLDER)
    db_path = Path(DATABASE_FILE)
    csv_path = Path(CSV_FILE)

    print("")
    print("============================================================")
    print("BUILDING PH DATABASE FROM Ver.03 EXCEL FILES ONLY")
    print("============================================================")
    print("pH folder:")
    print(f"  {ph_folder}")
    print("SQLite database:")
    print(f"  {db_path}")
    print("CSV output:")
    print(f"  {csv_path}")
    print("pH sheet:")
    print(f"  {PH_SHEET_NAME}")
    print("Filename filter:")
    print(f"  {FILE_NAME_FILTER}")
    print("Excel password if needed:")
    print(f"  {EXCEL_PASSWORD}")
    print("Columns:")
    print(f"  {PH_DUPLICATE_COL} = duplicate")
    print(f"  {PH_CODE_COL} = code")
    print(f"  {PH_VALUE_COL} = ph")
    print("Rows:")
    print("  27:47, 50:70, 73:93, ...")
    print("")
    print("Stop rules with OR logic:")
    print("  Stop current file if C is empty")
    print("  OR C has 3 or more '-' characters")
    print("  OR C does not contain a valid SU code")
    print("  OR H is not a meaningful numeric pH value")
    print("============================================================")

    if not ph_folder.exists():
        raise FileNotFoundError(f"pH folder not found: {ph_folder}")

    excel_files = get_ver03_excel_files(ph_folder)

    conn = create_database(db_path)

    csv_handle = open(csv_path, "w", newline="", encoding="utf-8-sig")
    writer = csv.writer(csv_handle)

    writer.writerow(
        [
            "duplicate",
            "code",
            "ph",
            "su_number",
            "source_file",
            "source_sheet",
            "source_row",
        ]
    )

    total_inserted = 0
    total_files_stopped = 0
    total_errors = 0

    for file_index, excel_file in enumerate(excel_files, start=1):
        print("")
        print("============================================================")
        print(f"[{file_index}/{len(excel_files)}] ACCESSING Ver.03 EXCEL FILE")
        print("============================================================")
        print("File:")
        print(f"  {excel_file}")

        try:
            print("")
            print("Opening workbook...")
            wb = open_excel_workbook(excel_file, password=EXCEL_PASSWORD)

            print("Workbook opened.")
            print("Available sheets:")
            for sheet in wb.sheetnames:
                print(f"  - {sheet}")

            if PH_SHEET_NAME not in wb.sheetnames:
                print("")
                print(f"SKIPPED FILE: sheet '{PH_SHEET_NAME}' not found.")
                wb.close()
                continue

            ws = wb[PH_SHEET_NAME]

            print("")
            print(f"Accessing sheet: {PH_SHEET_NAME}")
            print(f"Excel reported max row: {ws.max_row}")
            print("The script will stop when the first non-meaningful row is found.")

            file_inserted = 0
            stop_reason = ""

            # ------------------------------------------------------------
            # SAFE BLOCK LOOP
            # ------------------------------------------------------------
            # Blocks:
            #   27:47
            #   50:70
            #   73:93
            # etc.
            #
            # It stops the file immediately if the OR stop rule is triggered.
            # ------------------------------------------------------------

            block_start = PH_FIRST_ROW

            while True:
                block_end = block_start + BLOCK_SIZE - 1

                print("")
                print("------------------------------------------------------------")
                print(f"Reading pH block: rows {block_start} to {block_end}")
                print("------------------------------------------------------------")

                stop_this_file = False

                for row in range(block_start, block_end + 1):
                    duplicate_cell = f"{PH_DUPLICATE_COL}{row}"
                    code_cell = f"{PH_CODE_COL}{row}"
                    ph_cell = f"{PH_VALUE_COL}{row}"

                    raw_duplicate = ws[duplicate_cell].value
                    raw_code = ws[code_cell].value
                    raw_ph = ws[ph_cell].value

                    duplicate = normalize_duplicate(raw_duplicate)
                    code = normalize_code(raw_code)
                    ph_text = normalize_ph(raw_ph)
                    ph_number = parse_ph_number(raw_ph)
                    su_number = extract_su_number(code)

                    print("")
                    print(f"Reading row {row}:")
                    print(f"  {duplicate_cell} duplicate raw = {raw_duplicate!r} -> {duplicate!r}")
                    print(f"  {code_cell} code raw      = {raw_code!r} -> {code!r}")
                    print(f"  {ph_cell} ph raw          = {raw_ph!r} -> {ph_text!r}")
                    print(f"  pH numeric                = {ph_number}")
                    print(f"  Dash count in code        = {code.count('-')}")
                    print(f"  Extracted SU number       = {su_number}")

                    # ----------------------------------------------------
                    # OR STOP RULE
                    # ----------------------------------------------------
                    stop, reason = should_stop_file(raw_code, raw_ph)

                    if stop:
                        print("")
                        print("  STOP FILE:")
                        print(f"    Reason: {reason}")
                        print("    Moving to the next Excel file.")
                        stop_this_file = True
                        stop_reason = reason
                        total_files_stopped += 1
                        break

                    # At this point code and pH are both valid.
                    insert_record(
                        conn=conn,
                        duplicate=duplicate,
                        code=code,
                        ph=ph_number,
                        su_number=su_number,
                        source_file=excel_file.name,
                        source_sheet=PH_SHEET_NAME,
                        source_row=row,
                    )

                    writer.writerow(
                        [
                            duplicate,
                            code,
                            ph_number,
                            su_number,
                            excel_file.name,
                            PH_SHEET_NAME,
                            row,
                        ]
                    )

                    print("  INSERTED INTO DATABASE AND CSV:")
                    print(f"    duplicate = {duplicate}")
                    print(f"    code      = {code}")
                    print(f"    ph        = {ph_number}")
                    print(f"    file      = {excel_file.name}")
                    print(f"    row       = {row}")

                    file_inserted += 1

                if stop_this_file:
                    break

                # Move to next block:
                # 27:47 -> 50:70 -> 73:93
                block_start = block_end + GAP_ROWS + 1

                # Safety guard against formatted rows far below real data.
                if block_start > ws.max_row + BLOCK_SIZE + GAP_ROWS:
                    print("")
                    print("  SAFETY STOP:")
                    print("    Reached beyond Excel's reported max row.")
                    print("    Moving to next file.")
                    stop_reason = "Safety stop beyond Excel max row"
                    break

            conn.commit()
            wb.close()

            total_inserted += file_inserted

            print("")
            print("------------------------------------------------------------")
            print(f"Finished file: {excel_file.name}")
            print(f"Inserted rows: {file_inserted}")
            print(f"Stop reason:   {stop_reason}")
            print("------------------------------------------------------------")

        except Exception as e:
            total_errors += 1
            print("")
            print("ERROR while reading file:")
            print(f"  {excel_file}")
            print("Exception:")
            print(f"  {e}")

    csv_handle.close()

    cur = conn.cursor()
    cur.execute("SELECT COUNT(*) FROM ph_data")
    db_count = cur.fetchone()[0]

    conn.close()

    print("")
    print("============================================================")
    print("DATABASE BUILD FINISHED")
    print("============================================================")
    print(f"Total inserted rows: {total_inserted}")
    print(f"Rows in SQLite DB:   {db_count}")
    print(f"Files stopped by OR stop rule: {total_files_stopped}")
    print(f"Files with errors: {total_errors}")
    print("")
    print("Created:")
    print(f"  {db_path}")
    print(f"  {csv_path}")


if __name__ == "__main__":
    build_ph_database()
