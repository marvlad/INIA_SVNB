# update_colorimetric_report.py

from pathlib import Path
from shutil import copyfile
import re
import unicodedata
import warnings
from datetime import datetime
import argparse
import sys

from openpyxl import load_workbook
from openpyxl.drawing.image import Image


# ============================================================
# Windows UTF-8 console fix
# ============================================================

try:
    sys.stdout.reconfigure(encoding="utf-8")
    sys.stderr.reconfigure(encoding="utf-8")
except Exception:
    pass


warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")


# ============================================================
# Config
# ============================================================

TEMPLATE_DIR = "template"

TEMPLATES = {
    "bray": "F-82 Reporte de Resultados Colorimetricos Ver.05 Bray.xlsm",
    "olsen": "F-82 Reporte de Resultados Colorimetricos Ver.05 Olsen.xlsm",
}

INPUT_DAT = "input.dat"
CSV_DIR = "input"
OUTPUT_DIR = "output"

TARGET_SHEET = "P_DIS"

DATE_CELL = "T12"

# Excel columns in P_DIS for standards
CONCENTRATION_COLUMN = "F"
A882_COLUMN = "H"

START_ROW = 15

# Sample output location in P_DIS
SAMPLE_NAME_COLUMN = "C"
SAMPLE_A882_COLUMN = "K"

SAMPLE_START_ROW = 37
SAMPLES_PER_BLOCK = 20
ROWS_TO_SKIP = 3

IMAGE_DIR = "extracted_images_from_xlsm/media"


# ============================================================
# Command-line arguments
# ============================================================

def parse_args():
    parser = argparse.ArgumentParser(
        description="Generate analyzed colorimetric reports from CSV files."
    )

    parser.add_argument(
        "--method",
        choices=["Bray", "Olsen", "bray", "olsen"],
        required=True,
        help="Template method to use: Bray or Olsen",
    )

    return parser.parse_args()


def get_template_path(method):
    method_key = method.lower()

    if method_key not in TEMPLATES:
        raise ValueError(f"Unknown method: {method}")

    return Path(TEMPLATE_DIR) / TEMPLATES[method_key]


# ============================================================
# Helpers
# ============================================================

def norm(text):
    """
    Normalize text so accents and bad encoding do not break matching.
    """
    if text is None:
        return ""

    text = str(text).strip()
    text = text.replace("\x00", "")
    text = text.replace("??", "")

    text = unicodedata.normalize("NFKD", text)
    text = "".join(c for c in text if not unicodedata.combining(c))

    text = text.upper()
    text = re.sub(r"\s+", " ", text)

    return text.strip()


def parse_decimal(value):
    """
    Converts:
        0,05450 -> 0.05450
        1,00000 -> 1.00000
    """
    if value is None:
        return None

    value = str(value).strip()
    value = value.replace("\x00", "")

    if value == "":
        return None

    value = value.replace(",", ".")

    try:
        return float(value)
    except ValueError:
        return None


def round_3(value):
    """
    Round numeric values to exactly 3 decimals before writing to Excel.

    Example:
        0.05450 -> 0.055
        1.00000 -> 1.000
    """
    if value is None:
        return None

    return round(float(value), 3)


def read_text_auto_encoding(csv_file):
    """
    Reads CSV text trying common encodings.
    """
    csv_file = Path(csv_file)
    raw = csv_file.read_bytes()

    encodings = [
        "utf-16",
        "utf-16-le",
        "utf-16-be",
        "utf-8-sig",
        "latin-1",
    ]

    for enc in encodings:
        try:
            text = raw.decode(enc)

            if text.count("\x00") < 5:
                print(f"Detected CSV encoding for {csv_file.name}: {enc}")
                return text

        except UnicodeDecodeError:
            pass

    print(
        f"WARNING: Could not detect CSV encoding cleanly for "
        f"{csv_file.name}. Removing null bytes."
    )
    return raw.decode("latin-1", errors="ignore").replace("\x00", "")


def read_input_dat(input_dat):
    """
    Reads input.dat.
    """
    input_dat = Path(input_dat)

    if not input_dat.exists():
        raise FileNotFoundError(f"input.dat not found: {input_dat}")

    csv_files = []

    with open(input_dat, "r", encoding="utf-8-sig") as f:
        for line in f:
            line = line.strip()

            if not line:
                continue

            if line.startswith("#"):
                continue

            csv_files.append(line)

    if not csv_files:
        raise ValueError(f"No CSV files found in {input_dat}")

    print("\nCSV files listed in input.dat:")
    for csv_name in csv_files:
        print(f"  {csv_name}")

    return csv_files


def read_csv_rows(csv_file):
    """
    Reads CSV rows.

    Handles:
        - UTF-16
        - tab-separated
        - semicolon-separated
        - multiple-spaces-separated

    Important:
        We do not split by comma because decimals use comma.
    """
    text = read_text_auto_encoding(csv_file)
    text = text.replace("\x00", "")
    text = text.replace("??", "")

    rows = []

    for line in text.splitlines():
        line = line.strip()

        if not line:
            continue

        if "\t" in line:
            parts = line.split("\t")
        elif ";" in line:
            parts = line.split(";")
        elif re.search(r"\s{2,}", line):
            parts = re.split(r"\s{2,}", line)
        else:
            parts = [line]

        parts = [p.strip() for p in parts if p.strip() != ""]

        if parts:
            rows.append(parts)

    print("\nCSV rows detected:")
    for row in rows[:15]:
        print(row)

    return rows


def find_header_indices(rows):
    """
    Finds CSV columns:

        Nombre
        A882
        Concentracion

    Returns:
        sample_idx, a882_idx, concentration_idx
    """
    for row in rows:
        normalized = [norm(cell) for cell in row]

        sample_idx = None
        a882_idx = None
        concentration_idx = None

        for i, cell in enumerate(normalized):
            if "NOMBRE" in cell:
                sample_idx = i

            if "A882" in cell or cell == "882":
                a882_idx = i

            if "CONCENTRACION" in cell or cell.startswith("CONC"):
                concentration_idx = i

        if (
            sample_idx is not None
            and a882_idx is not None
            and concentration_idx is not None
        ):
            print("\nDetected CSV columns:")
            print(f"  Nombre column: {sample_idx}")
            print(f"  A882 column: {a882_idx}")
            print(f"  Concentracion column: {concentration_idx}")

            return sample_idx, a882_idx, concentration_idx

    raise ValueError("Could not find Nombre, A882 and Concentracion headers in CSV.")


def extract_date_from_filename(csv_file):
    """
    Extracts date from filename like:

        Cuantificación_23_03_2026_12_52_11.csv

    Returns:
        datetime.date
    """
    name = Path(csv_file).stem

    match = re.search(r"(\d{2})_(\d{2})_(\d{4})", name)

    if not match:
        raise ValueError(f"Could not extract date from filename: {csv_file}")

    day, month, year = match.groups()

    return datetime.strptime(f"{day}/{month}/{year}", "%d/%m/%Y").date()


def make_safe_filename(text):
    """
    Makes a safe filename from CSV stem.
    """
    text = str(text).strip()
    text = re.sub(r"[^\w\-]+", "_", text, flags=re.UNICODE)
    text = re.sub(r"_+", "_", text)
    return text.strip("_")


def convert_sample_name(raw_name):
    """
    Converts:

        SU 723 1 -> SU0723-ILL-26
        SU 62 1  -> SU0062-ILL-26

    It omits the last number after the space.

    If the name is not SU format, returns the original name.
    """
    if raw_name is None:
        return ""

    original = str(raw_name).strip()
    clean = norm(original)

    match = re.search(r"\bSU\s*0*(\d+)\s+\d+\b", clean)

    if match:
        su_number = int(match.group(1))
        return f"SU{su_number:04d}-ILL-26"

    return original


def get_sample_excel_row(
    sample_index,
    start_row=SAMPLE_START_ROW,
    samples_per_block=SAMPLES_PER_BLOCK,
    rows_to_skip=ROWS_TO_SKIP,
):
    """
    General row mapper.

    Starting from start_row, it writes samples_per_block rows,
    then skips rows_to_skip rows, and repeats.

    Example:
        37-56
        60-79
        83-102
        ...
    """
    block = sample_index // samples_per_block
    position = sample_index % samples_per_block

    return start_row + block * (samples_per_block + rows_to_skip) + position


# ============================================================
# CSV extraction
# ============================================================

def extract_standard_values_from_rows(rows, sample_idx, a882_idx, concentration_idx):
    """
    Extracts Estándar 1 to Estándar 7.

    Standards are written as:
        Concentracion -> F15:F21
        A882          -> H15:H21
    """
    standards = {}

    for row in rows:
        if len(row) <= max(sample_idx, a882_idx, concentration_idx):
            continue

        sample_name = norm(row[sample_idx])

        match = re.search(r"EST[A-Z]*NDAR\s*(\d+)", sample_name)

        if not match:
            continue

        standard_number = int(match.group(1))

        if 1 <= standard_number <= 7:
            a882_value = parse_decimal(row[a882_idx])
            concentration_value = parse_decimal(row[concentration_idx])

            if a882_value is None or concentration_value is None:
                print(f"WARNING: Could not parse standard row: {row}")
                continue

            standards[standard_number] = {
                "A882": a882_value,
                "Concentracion": concentration_value,
            }

            print(
                f"Found Estándar {standard_number}: "
                f"A882={a882_value:.3f}, "
                f"Concentracion={concentration_value:.3f}"
            )

    missing = [i for i in range(1, 8) if i not in standards]

    if missing:
        raise ValueError(f"Missing standards in CSV: {missing}")

    return standards


def extract_sample_values_from_rows(rows, sample_idx, a882_idx, concentration_idx):
    """
    Extracts all non-standard sample rows.

    It skips:
        - header row
        - Estándar 1 to Estándar 7

    Samples are written as:
        C37 = sample name
        K37 = A882

    Concentracion is still stored internally for printing/debugging,
    but it is not written to K.
    """
    samples = []

    for row in rows:
        if len(row) <= max(sample_idx, a882_idx, concentration_idx):
            continue

        raw_name = row[sample_idx].strip()
        clean_name = norm(raw_name)

        # Skip header row
        if "NOMBRE" in clean_name:
            continue

        # Skip standard rows
        if re.search(r"EST[A-Z]*NDAR\s*\d+", clean_name):
            continue

        a882_value = parse_decimal(row[a882_idx])
        concentration_value = parse_decimal(row[concentration_idx])

        if a882_value is None:
            print(f"WARNING: Could not parse sample A882 row: {row}")
            continue

        sample_name = convert_sample_name(raw_name)

        samples.append(
            {
                "raw_name": raw_name,
                "sample_name": sample_name,
                "A882": a882_value,
                "Concentracion": concentration_value,
            }
        )

        concentration_text = (
            f"{concentration_value:.3f}"
            if concentration_value is not None
            else "None"
        )

        print(
            f"Found sample: {raw_name} -> {sample_name}, "
            f"A882={a882_value:.3f}, "
            f"Concentracion={concentration_text}"
        )

    if not samples:
        raise ValueError("No sample rows found after standards.")

    return samples


def read_csv_data(csv_file):
    """
    Reads:
        - standards 1-7
        - all sample rows
    """
    rows = read_csv_rows(csv_file)
    sample_idx, a882_idx, concentration_idx = find_header_indices(rows)

    standards = extract_standard_values_from_rows(
        rows=rows,
        sample_idx=sample_idx,
        a882_idx=a882_idx,
        concentration_idx=concentration_idx,
    )

    samples = extract_sample_values_from_rows(
        rows=rows,
        sample_idx=sample_idx,
        a882_idx=a882_idx,
        concentration_idx=concentration_idx,
    )

    return standards, samples


# ============================================================
# Image insertion
# ============================================================

def reinsert_images(wb, image_dir):
    """
    Reinsert fixed images directly from:

        extracted_images_from_xlsm/media
    """
    image_dir = Path(image_dir)

    if not image_dir.exists():
        print(f"Image directory not found. Skipping image insertion: {image_dir}")
        return

    images_by_sheet = {
        "P_DIS": [
            ("image2.jpeg", "B1", 0.3588),
            ("image3.png", "B3", 0.5148),
            ("image4.png", "S20", 0.385),
        ],

        "Resultados": [
            ("image2.jpeg", "B1", 0.3588),
            ("image3.png", "B3", 0.5148),
        ],

        "Datos": [
            ("image2.jpeg", "B1", 0.3588),
            ("image3.png", "B3", 0.5148),
        ],
    }

    for sheet_name, images in images_by_sheet.items():
        if sheet_name not in wb.sheetnames:
            print(f"Sheet not found for image insertion: {sheet_name}")
            continue

        ws = wb[sheet_name]

        # Avoid duplicated or broken images
        ws._images = []

        for image_filename, anchor, scale in images:
            image_path = image_dir / image_filename

            if not image_path.exists():
                print(f"Image file not found: {image_path}")
                continue

            img = Image(str(image_path))
            img.width = img.width * scale
            img.height = img.height * scale

            ws.add_image(img, anchor)

            print(
                f"Inserted {image_filename} into {sheet_name}!{anchor} "
                f"at {int(scale * 100)}% size"
            )


# ============================================================
# Excel writing
# ============================================================

def write_standard_values(ws, standards):
    """
    Writes standards:

        CSV Concentracion -> Excel F15:F21
        CSV A882          -> Excel H15:H21

    Values are rounded to 3 decimals and displayed as 0.000.
    """
    for standard_number in range(1, 8):
        row = START_ROW + standard_number - 1

        concentration = round_3(standards[standard_number]["Concentracion"])
        a882 = round_3(standards[standard_number]["A882"])

        ws[f"{CONCENTRATION_COLUMN}{row}"] = concentration
        ws[f"{A882_COLUMN}{row}"] = a882

        ws[f"{CONCENTRATION_COLUMN}{row}"].number_format = "0.000"
        ws[f"{A882_COLUMN}{row}"].number_format = "0.000"


def clear_old_sample_values(ws, max_samples=300):
    """
    Clears old sample values using the same general block pattern.

    Clears:
        C = sample name
        K = A882
    """
    for i in range(max_samples):
        row = get_sample_excel_row(i)

        ws[f"{SAMPLE_NAME_COLUMN}{row}"] = None
        ws[f"{SAMPLE_A882_COLUMN}{row}"] = None


def write_sample_values(ws, samples):
    """
    Writes all samples in P_DIS using this general pattern:

        C37 = sample name
        K37 = A882

    A882 values are rounded to 3 decimals and displayed as 0.000.
    """
    clear_old_sample_values(ws)

    for i, sample in enumerate(samples):
        row = get_sample_excel_row(i)

        name_cell = f"{SAMPLE_NAME_COLUMN}{row}"
        a882_cell = f"{SAMPLE_A882_COLUMN}{row}"

        a882_value = round_3(sample["A882"])

        ws[name_cell] = sample["sample_name"]
        ws[a882_cell] = a882_value

        ws[a882_cell].number_format = "0.000"

        print(
            f"Written sample {i + 1}: "
            f"{name_cell} = {sample['sample_name']}, "
            f"{a882_cell} = {a882_value:.3f}"
        )


def create_output_path(csv_file, method):
    """
    Creates output path like:

        output/bray/Analizado_Bray_Cuantificacion_23_03_2026_12_52_11.xlsm
        output/olsen/Analizado_Olsen_Cuantificacion_23_03_2026_12_52_11.xlsm
    """
    csv_stem = make_safe_filename(Path(csv_file).stem)
    method_key = method.lower()
    method_name = method_key.capitalize()

    method_output_dir = Path(OUTPUT_DIR) / method_key
    method_output_dir.mkdir(parents=True, exist_ok=True)

    filename = f"Analizado_{method_name}_{csv_stem}.xlsm"

    return method_output_dir / filename


def update_report(input_xlsm, csv_file, image_dir, method):
    input_xlsm = Path(input_xlsm).resolve()
    csv_file = Path(csv_file).resolve()
    image_dir = Path(image_dir).resolve()

    if not input_xlsm.exists():
        raise FileNotFoundError(f"Template not found: {input_xlsm}")

    if not csv_file.exists():
        raise FileNotFoundError(f"CSV not found: {csv_file}")

    output_xlsm = create_output_path(csv_file, method).resolve()

    print("\n============================================================")
    print("Updating colorimetric report")
    print("============================================================")
    print(f"Method: {method.capitalize()}")
    print(f"Template: {input_xlsm}")
    print(f"CSV: {csv_file}")
    print(f"Images: {image_dir}")
    print(f"Output: {output_xlsm}")

    # Read CSV data
    date_value = extract_date_from_filename(csv_file)
    standards, samples = read_csv_data(csv_file)

    # Copy selected template once
    copyfile(input_xlsm, output_xlsm)

    # Open copied XLSM with macros preserved
    wb = load_workbook(output_xlsm, keep_vba=True)

    if TARGET_SHEET not in wb.sheetnames:
        print("Available sheets:")
        for sheet in wb.sheetnames:
            print(repr(sheet))

        raise KeyError(f"Target sheet not found: {TARGET_SHEET}")

    # Everything goes into P_DIS
    ws = wb[TARGET_SHEET]

    # Date in P_DIS
    ws[DATE_CELL] = date_value
    ws[DATE_CELL].number_format = "dd/mm/yyyy"

    # Standards 1-7 into P_DIS F15:F21 and H15:H21
    write_standard_values(ws, standards)

    # Samples into P_DIS C/K
    write_sample_values(ws, samples)

    # Reinsert fixed images
    reinsert_images(wb, image_dir=image_dir)

    # Save
    wb.save(output_xlsm)

    print("\nReport created successfully:")
    print(f"  {output_xlsm}")

    print("\nStandards written:")
    for standard_number in range(1, 8):
        row = START_ROW + standard_number - 1

        concentration = round_3(standards[standard_number]["Concentracion"])
        a882 = round_3(standards[standard_number]["A882"])

        print(
            f"  Estándar {standard_number}: "
            f"{CONCENTRATION_COLUMN}{row} = {concentration:.3f}, "
            f"{A882_COLUMN}{row} = {a882:.3f}"
        )

    print("\nSamples written in P_DIS:")
    for i, sample in enumerate(samples):
        row = get_sample_excel_row(i)
        a882 = round_3(sample["A882"])

        print(
            f"  Sample {i + 1}, row {row}: "
            f"{SAMPLE_NAME_COLUMN}{row} = {sample['sample_name']}, "
            f"{SAMPLE_A882_COLUMN}{row} = {a882:.3f}"
        )

    return output_xlsm


def main():
    args = parse_args()

    method = args.method.lower()

    input_xlsm = get_template_path(method)
    input_dat = Path(INPUT_DAT)
    csv_dir = Path(CSV_DIR)
    image_dir = Path(IMAGE_DIR)

    print("\nSelected method:")
    print(f"  {method.capitalize()}")

    print("\nSelected template:")
    print(f"  {input_xlsm}")

    print("\nOutput directory:")
    print(f"  {Path(OUTPUT_DIR) / method}")

    csv_names = read_input_dat(input_dat)

    created_reports = []
    failed_reports = []

    for csv_name in csv_names:
        csv_path = Path(csv_name)

        # If input.dat gives only filename, look inside CSV_DIR
        if not csv_path.is_absolute():
            csv_path = csv_dir / csv_path

        try:
            output_file = update_report(
                input_xlsm=input_xlsm,
                csv_file=csv_path,
                image_dir=image_dir,
                method=method,
            )
            created_reports.append(output_file)

        except Exception as e:
            print("\nERROR while processing:")
            print(f"  CSV: {csv_path}")
            print(f"  Error: {e}")
            failed_reports.append((csv_path, e))

    print("\n============================================================")
    print("Batch finished")
    print("============================================================")

    print("\nCreated reports:")
    for report in created_reports:
        print(f"  {report}")

    if failed_reports:
        print("\nFailed reports:")
        for csv_path, error in failed_reports:
            print(f"  {csv_path} -> {error}")
    else:
        print("\nNo failed reports.")


if __name__ == "__main__":
    main()
