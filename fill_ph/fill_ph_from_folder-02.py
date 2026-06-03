from pathlib import Path
import re
import shutil
import sqlite3
import argparse
from openpyxl import load_workbook


# ============================================================
# NORMALIZATION
# ============================================================

def normalize_text(value):
    if value is None:
        return ""

    text = str(value).strip()
    text = text.replace("\xa0", " ")
    text = re.sub(r"\s+", " ", text)
    text = text.lstrip("'").strip()

    return text


def normalize_code(value):
    return normalize_text(value).upper()


def extract_su_number(value):
    text = normalize_code(value)

    match = re.search(r"SU\s*0*(\d+)", text)

    if not match:
        return None

    return int(match.group(1))


def parse_ph(value):
    text = normalize_text(value)

    if text == "":
        return None

    text = text.replace(",", ".")

    try:
        return float(text)
    except Exception:
        return None


# ============================================================
# FILE NUMBER FILTER
# ============================================================

def extract_file_number(path):
    """
    Extract leading file number.

    Examples:
        23. F-82 Reporte...xlsx  -> 23
        24. F-82 Reporte...xlsx  -> 24
        34. F-82 Reporte...xlsx  -> 34
        ~$24. F-82...xlsx        -> None, skipped before this
    """
    name = Path(path).name.strip()

    match = re.match(r"^(\d+)", name)

    if not match:
        return None

    return int(match.group(1))


# ============================================================
# VERSION RULES FOR INPUT COLORIMETRIC FILES
# ============================================================

def detect_input_version(path):
    name = Path(path).name.upper()

    if "VER.04" in name or "VER 04" in name or "VER04" in name or "V04" in name:
        return "V04"

    if "VER.05" in name or "VER 05" in name or "VER05" in name or "V05" in name:
        return "V05"

    if "VER.02" in name or "VER 02" in name or "VER02" in name:
        return "V02"

    return "UNKNOWN"


def first_row_for_version(version):
    """
    User rule:
      V04 starts at C35 and pH goes to E35
      V05 starts at C37 and pH goes to E37
    """
    if version == "V04":
        return 35

    if version == "V05":
        return 37

    return 37


def block_settings_for_version(version):
    """
    V04 has 21 data rows per block:
      35-55
      58-78
      81-101

    V05 has 20 data rows per block:
      37-56
      60-79
      83-102
    """
    if version == "V04":
        return 21, 2

    if version == "V05":
        return 20, 3

    # fallback for older/unknown files
    return 20, 3


# ============================================================
# SQLITE LOOKUP
# ============================================================

def build_ph_lookup_from_sqlite(sqlite_path):
    sqlite_path = Path(sqlite_path)

    if not sqlite_path.exists():
        raise FileNotFoundError(f"SQLite database not found: {sqlite_path}")

    conn = sqlite3.connect(sqlite_path)
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()

    cur.execute(
        """
        SELECT
            code,
            ph,
            su_number,
            version,
            source_file,
            source_sheet,
            source_row
        FROM ph_data
        WHERE su_number IS NOT NULL
          AND ph IS NOT NULL
        """
    )

    rows = cur.fetchall()
    conn.close()

    lookup = {}

    for row in rows:
        code = normalize_code(row["code"])
        ph = parse_ph(row["ph"])

        try:
            su_number = int(row["su_number"])
        except Exception:
            su_number = extract_su_number(code)

        if su_number is None or ph is None:
            continue

        record = {
            "code": code,
            "ph": ph,
            "su_number": su_number,
            "db_version": row["version"],
            "source_file": row["source_file"],
            "source_sheet": row["source_sheet"],
            "source_row": row["source_row"],
        }

        old = lookup.get(su_number)

        # Keep largest pH if duplicate exists
        if old is None or ph > old["ph"]:
            lookup[su_number] = record

    print("")
    print("============================================================")
    print("pH DATABASE LOADED")
    print("============================================================")
    print(f"SQLite:      {sqlite_path}")
    print(f"Lookup size: {len(lookup)}")

    return lookup


# ============================================================
# FILE DISCOVERY
# ============================================================

def get_input_excel_files(input_dir, min_number=23, max_number=34):
    input_dir = Path(input_dir)

    files = sorted(
        list(input_dir.glob("*.xlsx")) +
        list(input_dir.glob("*.xlsm"))
    )

    selected = []

    for path in files:
        # Skip temporary or hidden Excel files
        if path.name.startswith("~"):
            print(f"Skipping temporary/hidden Excel file: {path.name}")
            continue

        file_number = extract_file_number(path)

        if file_number is None:
            print(f"Skipping file without leading number: {path.name}")
            continue

        if not (min_number <= file_number <= max_number):
            continue

        selected.append(path)

    print("")
    print("============================================================")
    print("INPUT FILE FILTER")
    print("============================================================")
    print(f"Input dir:      {input_dir}")
    print(f"Only numbers:   {min_number} to {max_number}")
    print(f"Files selected: {len(selected)}")
    for path in selected:
        print(f"  - {path.name}")

    return selected


# ============================================================
# SHEET HELPER
# ============================================================

def get_sheet(wb, preferred_sheet):
    if preferred_sheet in wb.sheetnames:
        return wb[preferred_sheet]

    raise ValueError(
        f"Sheet '{preferred_sheet}' not found. Available sheets: {wb.sheetnames}"
    )


# ============================================================
# ROW ITERATOR
# ============================================================

def iter_rows_by_blocks(ws, first_row, block_size, gap_rows, code_col, max_empty_blocks=3):
    """
    For V04:
      35-55
      58-78
      81-101

    For V05:
      37-56
      60-79
      83-102
    """
    block_start = first_row
    empty_blocks = 0

    while block_start <= ws.max_row:
        block_end = block_start + block_size - 1

        valid_codes = 0

        for row in range(block_start, min(block_end, ws.max_row) + 1):
            raw_code = ws[f"{code_col}{row}"].value
            su_number = extract_su_number(raw_code)

            if su_number is not None:
                valid_codes += 1

        if valid_codes == 0:
            empty_blocks += 1

            if empty_blocks >= max_empty_blocks:
                break

            block_start = block_end + gap_rows + 1
            continue

        empty_blocks = 0

        for row in range(block_start, min(block_end, ws.max_row) + 1):
            yield row

        block_start = block_end + gap_rows + 1


# ============================================================
# FILL ONE FILE
# ============================================================

def fill_one_file(
    input_file,
    output_file,
    ph_lookup,
    sheet_name="P_DIS",
    code_col="C",
    output_col="E",
    print_found=False,
):
    input_file = Path(input_file)
    output_file = Path(output_file)

    version = detect_input_version(input_file)
    first_row = first_row_for_version(version)
    block_size, gap_rows = block_settings_for_version(version)

    # IMPORTANT:
    # Inputs are not password protected.
    # keep_vba=True must be used ONLY for .xlsm.
    # For .xlsx, keep_vba=True can corrupt the output.
    keep_vba = input_file.suffix.lower() == ".xlsm"

    print("")
    print("============================================================")
    print("FILLING FILE")
    print("============================================================")
    print(f"Input:      {input_file}")
    print(f"Output:     {output_file}")
    print(f"Version:    {version}")
    print(f"Sheet:      {sheet_name}")
    print(f"Read:       {code_col}{first_row}")
    print(f"Write:      {output_col}{first_row}")
    print(f"Block size: {block_size}")
    print(f"Gap rows:   {gap_rows}")
    print(f"keep_vba:   {keep_vba}")

    output_file.parent.mkdir(parents=True, exist_ok=True)

    # Copy original file first, then write into the copy.
    shutil.copy2(input_file, output_file)

    # Open original only for reading values.
    wb_values = load_workbook(
        input_file,
        data_only=True,
        keep_vba=keep_vba,
    )
    ws_values = get_sheet(wb_values, sheet_name)

    # Open copied output for writing.
    wb_out = load_workbook(
        output_file,
        keep_vba=keep_vba,
    )
    ws_out = get_sheet(wb_out, sheet_name)

    written = 0
    not_found = 0
    skipped = 0

    for row in iter_rows_by_blocks(
        ws=ws_values,
        first_row=first_row,
        block_size=block_size,
        gap_rows=gap_rows,
        code_col=code_col,
    ):
        input_cell = f"{code_col}{row}"
        output_cell = f"{output_col}{row}"

        raw_code = ws_values[input_cell].value
        code = normalize_code(raw_code)
        su_number = extract_su_number(code)

        if code == "":
            skipped += 1
            continue

        if su_number is None:
            skipped += 1
            continue

        record = ph_lookup.get(su_number)

        if record is None:
            not_found += 1

            if print_found:
                print(
                    f"NOT FOUND: row={row} {input_cell}={code} "
                    f"SU={su_number}"
                )

            continue

        ph = record["ph"]
        ws_out[output_cell] = ph
        written += 1

        if print_found:
            print(
                f"WRITE: row={row} {input_cell}={code} "
                f"-> {output_cell}={ph} | "
                f"DB source_file={record.get('source_file', '')} | "
                f"source_sheet={record.get('source_sheet', '')} | "
                f"source_row={record.get('source_row', '')}"
            )

    wb_out.save(output_file)

    wb_values.close()
    wb_out.close()

    print("")
    print("Finished:")
    print(f"  Written:   {written}")
    print(f"  Not found: {not_found}")
    print(f"  Skipped:   {skipped}")

    return {
        "input": str(input_file),
        "output": str(output_file),
        "written": written,
        "not_found": not_found,
        "skipped": skipped,
    }


# ============================================================
# FILL FOLDER
# ============================================================

def fill_folder(
    input_dir,
    output_dir,
    ph_sqlite,
    min_number=23,
    max_number=34,
    sheet_name="P_DIS",
    code_col="C",
    output_col="E",
    print_found=False,
):
    input_dir = Path(input_dir)
    output_dir = Path(output_dir)
    ph_sqlite = Path(ph_sqlite)

    print("")
    print("============================================================")
    print("BATCH FILL pH FROM FOLDER")
    print("============================================================")
    print(f"Input dir:   {input_dir}")
    print(f"Output dir:  {output_dir}")
    print(f"SQLite DB:   {ph_sqlite}")
    print(f"Only files:  {min_number} to {max_number}")
    print("Rules:")
    print("  V04 -> read C35, write E35, block 21, gap 2")
    print("  V05 -> read C37, write E37, block 20, gap 3")
    print("  Skip files starting with ~")
    print("  .xlsx -> keep_vba=False")
    print("  .xlsm -> keep_vba=True")
    print("============================================================")

    if not input_dir.exists():
        raise FileNotFoundError(f"Input directory not found: {input_dir}")

    output_dir.mkdir(parents=True, exist_ok=True)

    ph_lookup = build_ph_lookup_from_sqlite(ph_sqlite)

    input_files = get_input_excel_files(
        input_dir=input_dir,
        min_number=min_number,
        max_number=max_number,
    )

    total_files = len(input_files)
    successful = 0
    failed = 0
    total_written = 0
    total_not_found = 0
    total_skipped = 0

    for index, input_file in enumerate(input_files, start=1):
        print("")
        print("############################################################")
        print(f"PROCESSING {index}/{total_files}")
        print("############################################################")

        output_file = output_dir / input_file.name

        try:
            result = fill_one_file(
                input_file=input_file,
                output_file=output_file,
                ph_lookup=ph_lookup,
                sheet_name=sheet_name,
                code_col=code_col,
                output_col=output_col,
                print_found=print_found,
            )

            successful += 1
            total_written += result["written"]
            total_not_found += result["not_found"]
            total_skipped += result["skipped"]

        except Exception as e:
            failed += 1
            print("")
            print("ERROR PROCESSING FILE:")
            print(f"  {input_file}")
            print(f"  {e}")

    print("")
    print("============================================================")
    print("BATCH FINISHED")
    print("============================================================")
    print(f"Files selected:   {total_files}")
    print(f"Successful:       {successful}")
    print(f"Failed:           {failed}")
    print(f"Total written:    {total_written}")
    print(f"Total not found:  {total_not_found}")
    print(f"Total skipped:    {total_skipped}")
    print(f"Output dir:       {output_dir}")


# ============================================================
# ARGUMENTS
# ============================================================

def parse_args():
    parser = argparse.ArgumentParser(
        description="Fill pH values into selected Excel files from a folder."
    )

    parser.add_argument(
        "--input-dir",
        default="/Users/mascenci/Downloads/Clientes",
        help="Folder containing input Excel files."
    )

    parser.add_argument(
        "--output-dir",
        default="/Users/mascenci/Downloads/Clientes_filled_ph",
        help="Folder where filled Excel files will be saved."
    )

    parser.add_argument(
        "--ph-sqlite",
        default="database2/ph_database.sqlite",
        help="Path to pH SQLite database."
    )

    parser.add_argument(
        "--min-number",
        type=int,
        default=23,
        help="First file number to process. Default: 23."
    )

    parser.add_argument(
        "--max-number",
        type=int,
        default=34,
        help="Last file number to process. Default: 34."
    )

    parser.add_argument(
        "--sheet-name",
        default="P_DIS",
        help="Sheet name to fill. Default: P_DIS."
    )

    parser.add_argument(
        "--code-col",
        default="C",
        help="Column containing SU codes. Default: C."
    )

    parser.add_argument(
        "--output-col",
        default="E",
        help="Column where pH is written. Default: E."
    )

    parser.add_argument(
        "--print-found",
        action="store_true",
        help="Print every pH written or not found, including DB source file/sheet/row."
    )

    return parser.parse_args()


def main():
    args = parse_args()

    fill_folder(
        input_dir=args.input_dir,
        output_dir=args.output_dir,
        ph_sqlite=args.ph_sqlite,
        min_number=args.min_number,
        max_number=args.max_number,
        sheet_name=args.sheet_name,
        code_col=args.code_col,
        output_col=args.output_col,
        print_found=args.print_found,
    )


if __name__ == "__main__":
    main()
