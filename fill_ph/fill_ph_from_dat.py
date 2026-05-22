# fill_ph_from_dat.py

from pathlib import Path
import re
import csv
import shutil
import argparse
from openpyxl import load_workbook


# ============================================================
# LOGGING
# ============================================================

def default_log(message):
    print(message)


# ============================================================
# NORMALIZATION
# ============================================================

def normalize_text(value):
    """
    Clean text.

    Handles:
        SU1149-ILL-26
        'SU1149-ILL-26
        spaces
        non-breaking spaces

    The leading Excel apostrophe is ignored.
    """
    if value is None:
        return ""

    text = str(value).strip()
    text = text.replace("\xa0", " ")
    text = re.sub(r"\s+", " ", text)

    text = text.lstrip("'").strip()

    return text


def normalize_code(value):
    return normalize_text(value).upper()


def extract_su_number(value):
    """
    Extract numeric SU number.

    Examples:
        SU1149-ILL-26   -> 1149
        'SU1149-ILL-26  -> 1149
        SU0079          -> 79
    """
    text = normalize_code(value)

    match = re.search(r"SU\s*0*(\d+)", text)

    if not match:
        return None

    return int(match.group(1))


def parse_ph(value):
    """
    Convert pH to float.

    Valid:
        7.5
        7,5
        "7.5"

    Invalid:
        ""
        "-"
        "---"
        "NO DATA"
    """
    text = normalize_text(value)

    if text == "":
        return None

    text = text.replace(",", ".")

    if set(text) <= {"-"}:
        return None

    try:
        return float(text)
    except ValueError:
        return None


# ============================================================
# CSV COLUMN HELPERS
# ============================================================

def find_column(fieldnames, possible_names):
    """
    Find a column using several possible names, case-insensitive.
    """
    normalized = {
        name.strip().lower(): name
        for name in fieldnames
    }

    for possible in possible_names:
        key = possible.strip().lower()
        if key in normalized:
            return normalized[key]

    return None


# ============================================================
# READ input_for_ph.dat
# ============================================================

def read_input_dat(input_dat_file, input_dir, log):
    """
    Read input_for_ph.dat.

    Each non-empty line should be an XLSM filename.

    Examples inside input_for_ph.dat:

        input1.xlsm
        input2.xlsm
        reports/input3.xlsm
        G:\\Mi unidad\\LABSAF ILLPA\\input4.xlsm

    If the path is relative, it is joined with input_dir.
    If the path is absolute, it is used directly.
    """
    input_dat_file = Path(input_dat_file)
    input_dir = Path(input_dir)

    log("")
    log("============================================================")
    log("READING input_for_ph.dat")
    log("============================================================")
    log(f"DAT file:  {input_dat_file}")
    log(f"Input dir: {input_dir}")

    if not input_dat_file.exists():
        raise FileNotFoundError(f"input_for_ph.dat not found: {input_dat_file}")

    files = []

    with open(input_dat_file, "r", encoding="utf-8-sig") as f:
        for line_number, line in enumerate(f, start=1):
            text = line.strip()

            if text == "":
                continue

            if text.startswith("#"):
                continue

            raw_path = Path(text)

            if raw_path.is_absolute():
                xlsm_path = raw_path
            else:
                xlsm_path = input_dir / raw_path

            files.append(xlsm_path)

            log(f"  Line {line_number}: {xlsm_path}")

    log("")
    log(f"Total input XLSM files listed: {len(files)}")

    return files


# ============================================================
# BUILD LOOKUP FROM CSV
# ============================================================

def build_ph_lookup_from_csv(csv_path, verbose=True, log_callback=None):
    """
    Read ph_database_Ver03.csv and build:

        su_number -> best record

    If the same SU appears more than once, keep the largest pH.
    """
    log = log_callback if log_callback is not None else default_log

    csv_path = Path(csv_path)

    log("")
    log("============================================================")
    log("READING pH CSV DATABASE")
    log("============================================================")
    log("CSV file:")
    log(f"  {csv_path}")

    if not csv_path.exists():
        raise FileNotFoundError(f"pH CSV not found: {csv_path}")

    lookup = {}

    total_rows = 0
    inserted_rows = 0
    skipped_rows = 0
    duplicate_replaced = 0
    duplicate_kept = 0

    with open(csv_path, "r", encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)

        if reader.fieldnames is None:
            raise ValueError("CSV file has no header row.")

        log("")
        log("Detected CSV columns:")
        for col in reader.fieldnames:
            log(f"  - {col}")

        code_col = find_column(
            reader.fieldnames,
            ["code", "su_code", "SU_CODE", "codigo", "código"]
        )

        ph_col = find_column(
            reader.fieldnames,
            ["ph", "PH", "ph_value", "PH_VALUE"]
        )

        source_file_col = find_column(
            reader.fieldnames,
            ["source_file", "SOURCE_FILE"]
        )

        source_row_col = find_column(
            reader.fieldnames,
            ["source_row", "SOURCE_ROW"]
        )

        if code_col is None:
            raise ValueError(
                "Could not find code column in CSV. "
                "Expected one of: code, su_code, SU_CODE, codigo, código."
            )

        if ph_col is None:
            raise ValueError(
                "Could not find pH column in CSV. "
                "Expected one of: ph, PH, ph_value, PH_VALUE."
            )

        log("")
        log(f"Using code column: {code_col}")
        log(f"Using pH column:   {ph_col}")

        for csv_row_number, row in enumerate(reader, start=2):
            total_rows += 1

            raw_code = row.get(code_col, "")
            raw_ph = row.get(ph_col, "")

            code = normalize_code(raw_code)
            su_number = extract_su_number(code)
            ph = parse_ph(raw_ph)

            source_file = row.get(source_file_col, "") if source_file_col else ""
            source_row = row.get(source_row_col, "") if source_row_col else ""

            if verbose:
                log("")
                log(f"CSV row {csv_row_number}:")
                log(f"  code raw = {raw_code!r} -> {code!r}")
                log(f"  SU number = {su_number}")
                log(f"  pH raw = {raw_ph!r} -> {ph}")
                log(f"  source file = {source_file}")
                log(f"  source row  = {source_row}")

            if code == "":
                if verbose:
                    log("  SKIPPED CSV ROW: empty code")
                skipped_rows += 1
                continue

            if su_number is None:
                if verbose:
                    log("  SKIPPED CSV ROW: no SU number")
                skipped_rows += 1
                continue

            if ph is None:
                if verbose:
                    log("  SKIPPED CSV ROW: invalid pH")
                skipped_rows += 1
                continue

            record = {
                "su_number": su_number,
                "code": code,
                "ph": ph,
                "source_file": source_file,
                "source_row": source_row,
                "csv_row": csv_row_number,
            }

            if su_number not in lookup:
                lookup[su_number] = record
                inserted_rows += 1

                if verbose:
                    log("  INSERTED INTO LOOKUP")

            else:
                old_record = lookup[su_number]

                if ph > old_record["ph"]:
                    lookup[su_number] = record
                    duplicate_replaced += 1

                    log("")
                    log("  DUPLICATE SU FOUND:")
                    log(f"    SU{su_number}")
                    log(f"    Old pH: {old_record['ph']} from {old_record['source_file']}")
                    log(f"    New pH: {ph} from {source_file}")
                    log("    Selected new value because it is larger.")

                else:
                    duplicate_kept += 1

                    if verbose:
                        log("")
                        log("  DUPLICATE SU FOUND:")
                        log(f"    SU{su_number}")
                        log(f"    Existing pH: {old_record['ph']} from {old_record['source_file']}")
                        log(f"    New pH: {ph} from {source_file}")
                        log("    Kept existing value because it is larger or equal.")

    log("")
    log("============================================================")
    log("CSV LOOKUP BUILD FINISHED")
    log("============================================================")
    log(f"Total CSV rows read:     {total_rows}")
    log(f"Inserted unique SUs:     {inserted_rows}")
    log(f"Skipped CSV rows:        {skipped_rows}")
    log(f"Duplicates replaced:     {duplicate_replaced}")
    log(f"Duplicates kept:         {duplicate_kept}")
    log(f"Final lookup size:       {len(lookup)}")

    return lookup


# ============================================================
# INPUT ROW ITERATOR
# ============================================================

def iter_input_rows(ws, first_row, block_size, gap_rows, log):
    """
    Generate input rows.

    Default:
        37 to 56
        60 to 79
        83 to 102
        106 to 125
        ...

    This ignores separator rows:
        36, 59, 82, 105, ...
    """
    row = first_row

    while row <= ws.max_row:
        block_start = row
        block_end = row + block_size - 1

        log("")
        log("------------------------------------------------------------")
        log(f"Input block: rows {block_start} to {block_end}")
        log("------------------------------------------------------------")

        for r in range(block_start, min(block_end, ws.max_row) + 1):
            yield r

        row = block_end + gap_rows + 1


# ============================================================
# SHEET HELPER
# ============================================================

def get_sheet(workbook, sheet_name):
    if sheet_name not in workbook.sheetnames:
        raise ValueError(
            f"Sheet '{sheet_name}' not found. "
            f"Available sheets: {workbook.sheetnames}"
        )

    return workbook[sheet_name]


# ============================================================
# FILL ONE XLSM
# ============================================================

def fill_one_xlsm_from_lookup(
    input_xlsm,
    output_xlsm,
    ph_lookup,
    input_sheet_name="P_DIS",
    first_row=37,
    block_size=20,
    gap_rows=3,
    input_code_col="C",
    input_output_col="E",
    verbose=True,
    log_callback=None,
):
    log = log_callback if log_callback is not None else default_log

    input_xlsm = Path(input_xlsm)
    output_xlsm = Path(output_xlsm)

    log("")
    log("============================================================")
    log("FILLING INPUT XLSM FROM pH CSV LOOKUP")
    log("============================================================")
    log("Input XLSM:")
    log(f"  {input_xlsm}")
    log("Input sheet:")
    log(f"  {input_sheet_name}")
    log("Output XLSM:")
    log(f"  {output_xlsm}")
    log("Reading input rows:")
    log(f"  {input_code_col}{first_row}:... by blocks")
    log("Writing pH to:")
    log(f"  {input_output_col}{first_row}:... by blocks")
    log("============================================================")

    if not input_xlsm.exists():
        raise FileNotFoundError(f"Input XLSM not found: {input_xlsm}")

    output_xlsm.parent.mkdir(parents=True, exist_ok=True)

    log("")
    log("Copying input XLSM to output XLSM...")
    shutil.copy2(input_xlsm, output_xlsm)
    log("Copy created:")
    log(f"  {output_xlsm}")

    log("")
    log("Opening input workbook for reading values...")
    wb_values = load_workbook(input_xlsm, data_only=True, keep_vba=True)
    ws_values = get_sheet(wb_values, input_sheet_name)

    log(f"Input sheet used: {ws_values.title}")
    log(f"Input max row: {ws_values.max_row}")

    log("")
    log("Opening output workbook for writing values...")
    wb_out = load_workbook(output_xlsm, keep_vba=True)
    ws_out = get_sheet(wb_out, input_sheet_name)

    written = 0
    not_found = 0
    skipped = 0

    detail_log = []

    for row in iter_input_rows(
        ws=ws_values,
        first_row=first_row,
        block_size=block_size,
        gap_rows=gap_rows,
        log=log,
    ):
        input_cell = f"{input_code_col}{row}"
        output_cell = f"{input_output_col}{row}"

        raw_code = ws_values[input_cell].value
        code = normalize_code(raw_code)
        su_number = extract_su_number(code)

        if verbose:
            log("")
            log("============================================================")
            log(f"INPUT ROW {row}")
            log("============================================================")
            log(f"Reading {input_cell}:")
            log(f"  Raw code:        {raw_code!r}")
            log(f"  Normalized code: {code!r}")
            log(f"  SU number:       {su_number}")

        if code == "":
            if verbose:
                log("  SKIPPED: empty input code")
            skipped += 1
            continue

        if su_number is None:
            if verbose:
                log("  SKIPPED: no SU number found")
            skipped += 1

            detail_log.append(
                {
                    "row": row,
                    "code": code,
                    "su_number": "",
                    "ph": "",
                    "status": "SKIPPED: no SU number",
                    "source_file": "",
                    "source_row": "",
                }
            )
            continue

        if su_number not in ph_lookup:
            if verbose:
                log("  NOT FOUND in pH CSV lookup")
            not_found += 1

            detail_log.append(
                {
                    "row": row,
                    "code": code,
                    "su_number": su_number,
                    "ph": "",
                    "status": "NOT FOUND",
                    "source_file": "",
                    "source_row": "",
                }
            )
            continue

        record = ph_lookup[su_number]
        ph = record["ph"]

        if verbose:
            log("  FOUND pH:")
            log(f"    pH:          {ph}")
            log(f"    CSV code:    {record['code']}")
            log(f"    Source file: {record['source_file']}")
            log(f"    Source row:  {record['source_row']}")
            log(f"    CSV row:     {record['csv_row']}")
            log("")
            log("  Writing to output workbook:")
            log(f"    Destination cell: {output_cell}")
            log(f"    Value: {ph}")

        ws_out[output_cell] = ph
        written += 1

        detail_log.append(
            {
                "row": row,
                "code": code,
                "su_number": su_number,
                "ph": ph,
                "status": "OK",
                "source_file": record["source_file"],
                "source_row": record["source_row"],
            }
        )

    log("")
    log("Saving output workbook...")
    wb_out.save(output_xlsm)

    wb_values.close()
    wb_out.close()

    log("")
    log("============================================================")
    log("FILLING FINISHED")
    log("============================================================")
    log(f"Output file created:")
    log(f"  {output_xlsm}")
    log(f"Written rows: {written}")
    log(f"Not found:    {not_found}")
    log(f"Skipped:      {skipped}")

    if verbose:
        log("")
        log("FINAL SUMMARY")
        log("============================================================")

        for item in detail_log:
            log(
                f"Row {item['row']:>4} | "
                f"C={item['code']:<25} | "
                f"SU={str(item['su_number']):<8} | "
                f"pH={str(item['ph']):<8} | "
                f"{item['status']:<20} | "
                f"{item['source_file']} | "
                f"source row={item['source_row']}"
            )

    return {
        "input_xlsm": str(input_xlsm),
        "output_xlsm": str(output_xlsm),
        "written": written,
        "not_found": not_found,
        "skipped": skipped,
    }


# ============================================================
# FILL MANY XLSM FILES FROM input_for_ph.dat
# ============================================================

def fill_many_xlsm_from_dat(
    input_dat_file,
    input_dir,
    ph_csv,
    output_dir,
    input_sheet_name="P_DIS",
    first_row=37,
    block_size=20,
    gap_rows=3,
    input_code_col="C",
    input_output_col="E",
    verbose=True,
    log_callback=None,
):
    log = log_callback if log_callback is not None else default_log

    input_dat_file = Path(input_dat_file)
    input_dir = Path(input_dir)
    ph_csv = Path(ph_csv)
    output_dir = Path(output_dir)

    log("")
    log("============================================================")
    log("BATCH FILL pH INTO XLSM FILES")
    log("============================================================")
    log(f"input_for_ph.dat: {input_dat_file}")
    log(f"Input dir:        {input_dir}")
    log(f"pH CSV:           {ph_csv}")
    log(f"Output dir:       {output_dir}")
    log("Output rule:")
    log("  output_dir / same input filename")
    log("============================================================")

    output_dir.mkdir(parents=True, exist_ok=True)

    ph_lookup = build_ph_lookup_from_csv(
        csv_path=ph_csv,
        verbose=verbose,
        log_callback=log,
    )

    input_files = read_input_dat(
        input_dat_file=input_dat_file,
        input_dir=input_dir,
        log=log,
    )

    total_files = len(input_files)
    successful_files = 0
    failed_files = 0
    total_written = 0
    total_not_found = 0
    total_skipped = 0

    file_results = []

    for index, input_xlsm in enumerate(input_files, start=1):
        log("")
        log("############################################################")
        log(f"PROCESSING FILE {index}/{total_files}")
        log("############################################################")
        log(f"Input file: {input_xlsm}")

        output_xlsm = output_dir / input_xlsm.name

        try:
            result = fill_one_xlsm_from_lookup(
                input_xlsm=input_xlsm,
                output_xlsm=output_xlsm,
                ph_lookup=ph_lookup,
                input_sheet_name=input_sheet_name,
                first_row=first_row,
                block_size=block_size,
                gap_rows=gap_rows,
                input_code_col=input_code_col,
                input_output_col=input_output_col,
                verbose=verbose,
                log_callback=log,
            )

            successful_files += 1
            total_written += result["written"]
            total_not_found += result["not_found"]
            total_skipped += result["skipped"]

            file_results.append(
                {
                    "input": str(input_xlsm),
                    "output": str(output_xlsm),
                    "status": "OK",
                    "written": result["written"],
                    "not_found": result["not_found"],
                    "skipped": result["skipped"],
                    "error": "",
                }
            )

        except Exception as e:
            failed_files += 1

            log("")
            log("ERROR PROCESSING FILE:")
            log(f"  {input_xlsm}")
            log(f"  {e}")

            file_results.append(
                {
                    "input": str(input_xlsm),
                    "output": str(output_xlsm),
                    "status": "ERROR",
                    "written": 0,
                    "not_found": 0,
                    "skipped": 0,
                    "error": str(e),
                }
            )

    log("")
    log("============================================================")
    log("BATCH FINISHED")
    log("============================================================")
    log(f"Total files listed:   {total_files}")
    log(f"Successful files:     {successful_files}")
    log(f"Failed files:         {failed_files}")
    log(f"Total pH written:     {total_written}")
    log(f"Total not found:      {total_not_found}")
    log(f"Total skipped:        {total_skipped}")
    log(f"Output directory:     {output_dir}")

    return {
        "total_files": total_files,
        "successful_files": successful_files,
        "failed_files": failed_files,
        "total_written": total_written,
        "total_not_found": total_not_found,
        "total_skipped": total_skipped,
        "output_dir": str(output_dir),
        "file_results": file_results,
    }


# ============================================================
# ARGUMENTS
# ============================================================

def parse_args():
    parser = argparse.ArgumentParser(
        description="Fill pH values into many XLSM files listed in input_for_ph.dat."
    )

    parser.add_argument(
        "--input-dat",
        required=True,
        help="Path to input_for_ph.dat containing the list of XLSM filenames."
    )

    parser.add_argument(
        "--input-dir",
        required=True,
        help="Base folder where the XLSM files are located."
    )

    parser.add_argument(
        "--ph-csv",
        required=True,
        help="Path to ph_database_Ver03.csv."
    )

    parser.add_argument(
        "--output-dir",
        required=True,
        help="Folder where output XLSM files will be saved."
    )

    parser.add_argument(
        "--sheet-name",
        default="P_DIS",
        help="Input workbook sheet name. Default: P_DIS"
    )

    parser.add_argument(
        "--first-row",
        type=int,
        default=37,
        help="First row to read in each block. Default: 37"
    )

    parser.add_argument(
        "--block-size",
        type=int,
        default=20,
        help="Number of rows per block. Default: 20"
    )

    parser.add_argument(
        "--gap-rows",
        type=int,
        default=3,
        help="Number of rows between blocks. Default: 3"
    )

    parser.add_argument(
        "--code-col",
        default="C",
        help="Column containing SU codes. Default: C"
    )

    parser.add_argument(
        "--output-col",
        default="E",
        help="Column where pH will be written. Default: E"
    )

    parser.add_argument(
        "--quiet",
        action="store_true",
        help="Reduce row-by-row logging."
    )

    return parser.parse_args()


def main():
    args = parse_args()

    fill_many_xlsm_from_dat(
        input_dat_file=args.input_dat,
        input_dir=args.input_dir,
        ph_csv=args.ph_csv,
        output_dir=args.output_dir,
        input_sheet_name=args.sheet_name,
        first_row=args.first_row,
        block_size=args.block_size,
        gap_rows=args.gap_rows,
        input_code_col=args.code_col,
        input_output_col=args.output_col,
        verbose=not args.quiet,
    )


if __name__ == "__main__":
    main()
